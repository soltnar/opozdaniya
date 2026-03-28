#!/usr/bin/env python3
from __future__ import annotations

import json
import io
import math
import re
import statistics
import subprocess
import sys
import threading
import traceback
import uuid
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

ROOT = Path(__file__).resolve().parent
WEB_DIR = ROOT / "web"
EXPORT_SCRIPT = ROOT / "export_delivery_statuses.py"
HOST = "127.0.0.1"
PORT = 8765

JOB_LOCK = threading.Lock()
JOBS: dict[str, dict] = {}
LATEST_JOB_ID: str | None = None
RUNNING_PROCESSES: dict[str, subprocess.Popen] = {}


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def append_log(job_id: str, line: str) -> None:
    with JOB_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return
        logs = job.setdefault("logs", [])
        logs.append(line.rstrip("\n"))
        if len(logs) > 5000:
            del logs[:1000]


def update_job(job_id: str, **kwargs) -> None:
    with JOB_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return
        job.update(kwargs)


def run_export_job(job_id: str, date_value: str) -> None:
    def detect_har_path() -> Path | None:
        names = [
            "rest.saby.ru_dost.har",
            "rest.saby.ru history.har",
            "rest.saby.ru_history.har",
        ]
        roots = [
            ROOT,
            Path.cwd(),
            Path.home(),
            Path.home() / "Desktop",
            Path.home() / "Downloads",
        ]
        seen: set[str] = set()
        for root in roots:
            try:
                root = root.resolve()
            except Exception:
                continue
            key = str(root)
            if key in seen:
                continue
            seen.add(key)
            for name in names:
                p = (root / name)
                if p.exists() and p.is_file():
                    return p.resolve()
        return None

    har_path = detect_har_path()

    if getattr(sys, "frozen", False):
        base_cmd = [
            str(Path(sys.executable).resolve()),
            "--worker",
            "--date",
            date_value,
        ]
    else:
        worker_exe = Path(sys.executable).resolve().with_name("export_delivery_statuses.exe")
        if worker_exe.exists():
            base_cmd = [
                str(worker_exe),
                "--date",
                date_value,
            ]
        else:
            base_cmd = [
                sys.executable,
                "-u",
                str(EXPORT_SCRIPT),
                "--date",
                date_value,
            ]
    if har_path is not None:
        base_cmd.extend(["--har", str(har_path)])
        append_log(job_id, f"[server] HAR auto-detected: {har_path}")
    output_path = None

    def run_once(cmd: list[str]) -> int:
        nonlocal output_path
        append_log(job_id, f"[server] start: {' '.join(cmd)}")

        process = subprocess.Popen(
            cmd,
            cwd=str(ROOT),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
        )
        with JOB_LOCK:
            RUNNING_PROCESSES[job_id] = process
            job = JOBS.get(job_id)
            if job is not None:
                job["pid"] = process.pid

        assert process.stdout is not None
        for line in process.stdout:
            append_log(job_id, line)
            if "Excel сохранен:" in line:
                output_path = line.split("Excel сохранен:", 1)[1].strip()

        return_code = process.wait()
        with JOB_LOCK:
            RUNNING_PROCESSES.pop(job_id, None)
        return return_code

    try:
        return_code = run_once(base_cmd)
        with JOB_LOCK:
            stop_requested = bool(JOBS.get(job_id, {}).get("stop_requested"))
        if stop_requested:
            update_job(
                job_id,
                status="stopped",
                ended_at=now_iso(),
                error=None,
            )
            append_log(job_id, "[server] stopped by user")
            return

        if return_code == 0:
            update_job(
                job_id,
                status="success",
                ended_at=now_iso(),
                output_path=output_path,
            )
            append_log(job_id, "[server] completed successfully")
        else:
            update_job(
                job_id,
                status="error",
                ended_at=now_iso(),
                error=f"process exited with code {return_code}",
            )
            append_log(job_id, f"[server] failed with code {return_code}")
    except Exception as err:  # noqa: BLE001
        with JOB_LOCK:
            RUNNING_PROCESSES.pop(job_id, None)
        update_job(
            job_id,
            status="error",
            ended_at=now_iso(),
            error=str(err),
        )
        append_log(job_id, f"[server] exception: {err}")
        append_log(job_id, traceback.format_exc())


def create_job(date_value: str) -> dict:
    global LATEST_JOB_ID

    with JOB_LOCK:
        for job in JOBS.values():
            if job.get("status") == "running":
                raise RuntimeError("Уже есть активная задача. Дождитесь завершения.")

        job_id = uuid.uuid4().hex[:12]
        JOBS[job_id] = {
            "id": job_id,
            "date": date_value,
            "status": "running",
            "started_at": now_iso(),
            "ended_at": None,
            "logs": [],
            "output_path": None,
            "error": None,
            "pid": None,
            "stop_requested": False,
        }
        LATEST_JOB_ID = job_id

    thread = threading.Thread(target=run_export_job, args=(job_id, date_value), daemon=True)
    thread.start()

    return JOBS[job_id]


def stop_job(job_id: str | None = None) -> dict:
    with JOB_LOCK:
        target_id = job_id or LATEST_JOB_ID
        if not target_id or target_id not in JOBS:
            raise RuntimeError("Задача не найдена.")
        job = JOBS[target_id]
        if job.get("status") != "running":
            raise RuntimeError("Активная задача не запущена.")
        process = RUNNING_PROCESSES.get(target_id)
        if process is None:
            raise RuntimeError("Процесс задачи не найден.")
        job["stop_requested"] = True

    append_log(target_id, "[server] stop requested by user")
    try:
        process.terminate()
        process.wait(timeout=5)
    except Exception:  # noqa: BLE001
        try:
            process.kill()
        except Exception:  # noqa: BLE001
            pass

    with JOB_LOCK:
        job = JOBS.get(target_id, {})
    return {"job_id": target_id, "status": job.get("status", "stopping")}


def run_embedded_worker(date_value: str) -> int:
    # Worker mode for frozen single-exe build: execute export script inside this process.
    import export_delivery_statuses  # noqa: PLC0415

    saved_argv = list(sys.argv)
    sys.argv = ["export_delivery_statuses.py", "--date", date_value]
    try:
        try:
            result = export_delivery_statuses.main()
        except SystemExit as exc:
            code = exc.code
            if isinstance(code, int):
                return code
            return 1 if code else 0
        return int(result) if isinstance(result, int) else 0
    finally:
        sys.argv = saved_argv


def _safe_float(value):
    if value in (None, ""):
        return None
    try:
        num = float(value)
        if math.isnan(num) or math.isinf(num):
            return None
        return num
    except (TypeError, ValueError):
        return None


def _percentile(values: list[float], q: float):
    if not values:
        return None
    ordered = sorted(values)
    idx = max(0, min(len(ordered) - 1, math.ceil(q * len(ordered)) - 1))
    return ordered[idx]


def _sheet_dict_rows(ws) -> list[dict]:
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []
    norm_headers = [str(h).strip() if h is not None else "" for h in headers]
    out: list[dict] = []
    for values in rows_iter:
        if values is None:
            continue
        if all(v in (None, "") for v in values):
            continue
        row: dict = {}
        for i, key in enumerate(norm_headers):
            if not key:
                continue
            row[key] = values[i] if i < len(values) else None
        out.append(row)
    return out


def _resolve_job_output(job_id: str | None):
    with JOB_LOCK:
        target_id = job_id or LATEST_JOB_ID
        job = JOBS.get(target_id) if target_id else None

    if not job:
        raise RuntimeError("job not found")
    if job.get("status") != "success":
        raise RuntimeError("job is not completed successfully")
    output_path_raw = job.get("output_path")
    if not output_path_raw:
        raise RuntimeError("output file is missing")

    file_path = Path(str(output_path_raw)).expanduser().resolve()
    allowed_root = (ROOT / "exports").resolve()
    if not str(file_path).startswith(str(allowed_root)):
        raise RuntimeError("forbidden output path")
    if not file_path.exists() or not file_path.is_file():
        raise RuntimeError("output file not found")
    return target_id, job, file_path


def _resolve_output_by_date(date_value: str):
    try:
        datetime.strptime(date_value, "%Y-%m-%d")
    except ValueError as exc:
        raise RuntimeError("date must be YYYY-MM-DD") from exc

    file_path = (ROOT / "exports" / f"order_status_history_{date_value}.xlsx").resolve()
    allowed_root = (ROOT / "exports").resolve()
    if not str(file_path).startswith(str(allowed_root)):
        raise RuntimeError("forbidden output path")
    if not file_path.exists() or not file_path.is_file():
        raise RuntimeError("output file not found")
    return file_path


def _sort_orders(rows: list[dict], sort_mode: str | None) -> list[dict]:
    mode = (sort_mode or "restaurant_asc").strip().lower()

    def _flt(value: Any, default: float = -1e12) -> float:
        try:
            v = float(value)
            if math.isnan(v) or math.isinf(v):
                return default
            return v
        except Exception:
            return default

    def _dt(value: Any) -> datetime | None:
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value
        s = str(value).strip().replace("T", " ").replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(s)
        except Exception:
            return None

    if mode == "restaurant_desc":
        return sorted(
            rows,
            key=lambda x: (
                str(x.get("restaurant") or "").lower(),
                _flt(x.get("total_min"), default=-1.0),
                str(x.get("number") or x.get("sale") or ""),
            ),
            reverse=True,
        )
    if mode == "total_desc":
        return sorted(
            rows,
            key=lambda x: (
                _flt(x.get("total_min"), default=-1.0),
                str(x.get("restaurant") or "").lower(),
            ),
            reverse=True,
        )
    if mode == "promised_delta_desc":
        return sorted(
            rows,
            key=lambda x: (
                _flt(x.get("promised_delta_min"), default=-1.0),
                _flt(x.get("total_min"), default=-1.0),
                str(x.get("restaurant") or "").lower(),
            ),
            reverse=True,
        )
    if mode == "promised_time_asc":
        return sorted(
            rows,
            key=lambda x: (
                _dt(x.get("promised_time")) or datetime.max,
                str(x.get("restaurant") or "").lower(),
                str(x.get("number") or x.get("sale") or ""),
            ),
        )
    return sorted(
        rows,
        key=lambda x: (
            str(x.get("restaurant") or "").lower(),
            _flt(x.get("promised_delta_min"), default=-1.0),
            _flt(x.get("total_min"), default=-1.0),
            str(x.get("number") or x.get("sale") or ""),
        ),
    )


def build_analytics_payload(
    file_path: Path,
    restaurant_filter: str | None = None,
    sort_mode: str | None = None,
) -> dict:
    wb = load_workbook(filename=str(file_path), data_only=True, read_only=True)

    def parse_dt(value):
        def _naive(dt_value: datetime) -> datetime:
            if dt_value.tzinfo is not None:
                return dt_value.astimezone().replace(tzinfo=None)
            return dt_value

        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return _naive(value)
        text = str(value).strip().replace("Z", "+00:00")
        text = re.sub(r"([+-]\d{2})$", r"\1:00", text)
        text = text.replace("T", " ")
        try:
            return _naive(datetime.fromisoformat(text))
        except ValueError:
            for fmt in (
                "%Y-%m-%d %H:%M:%S.%f",
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d %H:%M",
                "%d.%m.%Y %H:%M:%S",
                "%d.%m.%Y %H:%M",
            ):
                try:
                    return _naive(datetime.strptime(text, fmt))
                except ValueError:
                    continue
        return None

    def norm_status(value: Any) -> str:
        if value in (None, ""):
            return ""
        return str(value).strip().lower().replace("ё", "е")

    registry_rows = _sheet_dict_rows(wb["Реестр"]) if "Реестр" in wb.sheetnames else []
    status_rows = _sheet_dict_rows(wb["Статусы"]) if "Статусы" in wb.sheetnames else []

    if not registry_rows and not status_rows:
        return {
            "kpi": {
                "orders": 0,
                "avg_total_min": None,
                "p90_total_min": None,
                "avg_delivery_min": None,
                "p90_delivery_min": None,
                "overdue_count": 0,
                "overdue_rate": 0.0,
                "delivery_orders": 0,
                "pickup_orders": 0,
                "orders_with_statuses": 0,
                "no_delivery_stage_count": 0,
                "no_delivery_stage_rate": 0.0,
            },
            "thresholds": {
                "overdue_total_min": 60.0,
                "late_delivery_min": 60.0,
                "critical_delivery_min": 90.0,
            },
            "stages": [],
            "bottlenecks": [],
            "hotspots": [],
            "restaurant_totals": [],
            "problem_orders": [],
            "status_flow": {
                "orders_with_statuses": 0,
                "no_delivery_stage_count": 0,
                "no_delivery_stage_rate": 0.0,
                "transitions": [],
                "phase_stats": [],
            },
            "load_by_hour": [],
            "orders": [],
            "notice": "Нет данных для построения аналитики.",
        }

    orders_meta: dict[Any, dict] = {}
    for row in registry_rows:
        sale = row.get("Sale")
        if sale in (None, ""):
            continue
        restaurant = (
            row.get("WarehouseName")
            or row.get("CompanyName")
            or row.get("RealCompanyName")
            or row.get("OriginCompanyName")
            or "Не указан"
        )
        orders_meta[sale] = {
            "sale": sale,
            "number": row.get("Number"),
            "restaurant": str(restaurant),
            "courier": row.get("CourierName"),
            "operator": row.get("OperatorName") or row.get("EmployeeName"),
            "order_date": row.get("DateWTZ"),
            "closed_date": row.get("ClosedWTZ"),
            "promised_date": row.get("NextDateWTZText"),
            "source": row.get("Source"),
        }

    events_by_sale: dict[Any, list[tuple[datetime, str, str]]] = defaultdict(list)
    transitions_counter: Counter[tuple[str, str]] = Counter()
    for row in status_rows:
        sale = row.get("Sale")
        if sale in (None, ""):
            continue
        status_from = str(row.get("StatusFrom") or "").strip()
        status_to = str(row.get("StatusTo") or "").strip()
        if status_from or status_to:
            transitions_counter[(status_from, status_to)] += 1
        tm = parse_dt(row.get("StatusTime"))
        if tm is None:
            continue
        events_by_sale[sale].append((tm, status_from, status_to))

    overdue_total_threshold = 60.0
    late_delivery_threshold = 60.0
    critical_delivery_threshold = 90.0

    detailed_orders: list[dict] = []
    processing_vals: list[float] = []
    cooking_vals: list[float] = []
    assembly_vals: list[float] = []
    delivery_vals: list[float] = []
    total_vals: list[float] = []
    bottlenecks_counter: Counter[str] = Counter()
    phase_values_map: dict[str, list[float]] = {
        "Не распределен->Выполнение": [],
        "Выполнение->Сборка/Доставка": [],
        "Сборка->Доставка/Самовывоз": [],
        "Доставка->Выполнен": [],
    }

    all_sales = set(orders_meta.keys()) | set(events_by_sale.keys())

    for sale in all_sales:
        meta = orders_meta.get(sale, {})
        events = sorted(events_by_sale.get(sale, []), key=lambda x: x[0])
        is_pickup = any(
            "самовывоз" in f"{norm_status(frm)} {norm_status(to)}" for _, frm, to in events
        )
        promised_time = parse_dt(meta.get("promised_date"))
        sla_start_anchor = (promised_time - timedelta(hours=1)) if promised_time else None

        processing_min = None
        cooking_min = None
        assembly_min = None
        delivery_min = None
        pickup_wait_min = None

        processing_acc = 0.0
        cooking_acc = 0.0
        assembly_acc = 0.0
        delivery_acc = 0.0
        pickup_wait_acc = 0.0

        for idx in range(1, len(events)):
            prev_time, _, prev_to = events[idx - 1]
            cur_time, _, cur_to = events[idx]
            prev_to_norm = norm_status(prev_to)
            cur_to_norm = norm_status(cur_to)
            delta = (cur_time - prev_time).total_seconds() / 60.0
            if delta < 0 or delta > 24 * 60:
                continue

            if prev_to_norm in {"не распределен", "не выбран"} and cur_to_norm == "выполнение":
                processing_acc += delta
                phase_values_map["Не распределен->Выполнение"].append(delta)
            if prev_to_norm == "выполнение" and cur_to_norm in {"сборка", "доставка"}:
                cooking_delta = delta
                # Для заказов "ко времени" готовку считаем не ранее, чем за 60 минут до планового времени.
                if sla_start_anchor is not None and prev_time < sla_start_anchor:
                    if cur_time <= sla_start_anchor:
                        cooking_delta = 0.0
                    else:
                        cooking_delta = (cur_time - sla_start_anchor).total_seconds() / 60.0
                if cooking_delta > 0:
                    cooking_acc += cooking_delta
                    phase_values_map["Выполнение->Сборка/Доставка"].append(cooking_delta)
            if prev_to_norm == "сборка" and cur_to_norm in {"доставка", "самовывоз"}:
                assembly_acc += delta
                phase_values_map["Сборка->Доставка/Самовывоз"].append(delta)
            if prev_to_norm == "доставка" and cur_to_norm == "выполнен":
                delivery_delta = delta
                # Для заказов "ко времени" доставку также считаем не ранее T-60.
                if sla_start_anchor is not None and prev_time < sla_start_anchor:
                    if cur_time <= sla_start_anchor:
                        delivery_delta = 0.0
                    else:
                        delivery_delta = (cur_time - sla_start_anchor).total_seconds() / 60.0
                if delivery_delta > 0:
                    delivery_acc += delivery_delta
                    phase_values_map["Доставка->Выполнен"].append(delivery_delta)
            if prev_to_norm == "самовывоз" and cur_to_norm == "выполнен":
                pickup_wait_acc += delta

        if processing_acc > 0:
            processing_min = processing_acc
            processing_vals.append(processing_min)
        if cooking_acc > 0:
            cooking_min = cooking_acc
            cooking_vals.append(cooking_min)
        if assembly_acc > 0:
            assembly_min = assembly_acc
            assembly_vals.append(assembly_min)
        if delivery_acc > 0:
            delivery_min = delivery_acc
            delivery_vals.append(delivery_min)
        if pickup_wait_acc > 0:
            pickup_wait_min = pickup_wait_acc

        start_time = events[0][0] if events else parse_dt(meta.get("order_date"))
        pickup_ready_time = None
        done_time = None
        for tm, _, status_to in events:
            if norm_status(status_to) == "самовывоз" and pickup_ready_time is None:
                pickup_ready_time = tm
            if norm_status(status_to) == "выполнен":
                done_time = tm
        if done_time is None:
            done_time = parse_dt(meta.get("closed_date"))

        total_min = None
        if start_time is not None and done_time is not None:
            effective_start_time = start_time
            # Для заказов "ко времени" не учитываем ожидание с прошлого дня до T-60.
            if sla_start_anchor is not None and effective_start_time < sla_start_anchor:
                effective_start_time = sla_start_anchor
            total_end_time = done_time
            # Для самовывоза SLA считаем до готовности к выдаче, а не до фактического закрытия.
            if is_pickup and pickup_ready_time is not None:
                total_end_time = pickup_ready_time
            delta_total = (total_end_time - effective_start_time).total_seconds() / 60.0
            if 0 <= delta_total <= 24 * 60:
                total_min = delta_total
                total_vals.append(total_min)

        promised_delta_min = None
        if promised_time is not None and done_time is not None:
            promised_target_time = done_time
            if is_pickup and pickup_ready_time is not None:
                promised_target_time = pickup_ready_time
            promised_delta_min = (promised_target_time - promised_time).total_seconds() / 60.0

        stage_for_bottleneck = {}
        if processing_min is not None:
            stage_for_bottleneck["Обработка"] = processing_min
        if cooking_min is not None:
            stage_for_bottleneck["Готовка"] = cooking_min
        if assembly_min is not None:
            stage_for_bottleneck["Сборка"] = assembly_min
        if delivery_min is not None:
            stage_for_bottleneck["Доставка"] = delivery_min
        if is_pickup and pickup_wait_min is not None:
            stage_for_bottleneck["Выдача самовывоза"] = pickup_wait_min

        bottleneck_stage = None
        bottleneck_min = None
        if stage_for_bottleneck:
            bottleneck_stage, bottleneck_min = max(stage_for_bottleneck.items(), key=lambda x: x[1])
            bottlenecks_counter[bottleneck_stage] += 1
        else:
            bottlenecks_counter["Не определено"] += 1

        missing_stages: list[str] = []
        if is_pickup:
            if pickup_ready_time is None:
                missing_stages.append("нет этапа Самовывоз")
        else:
            if processing_min is None:
                missing_stages.append("нет этапа Обработка")
            if cooking_min is None:
                missing_stages.append("нет этапа Готовка")
            if assembly_min is None:
                missing_stages.append("нет этапа Сборка")
            if delivery_min is None:
                missing_stages.append("нет этапа Доставка")

        overdue = bool(total_min is not None and total_min > overdue_total_threshold)
        reason_parts: list[str] = []
        if bottleneck_stage and bottleneck_min is not None:
            reason_parts.append(f"{bottleneck_stage}: {bottleneck_min:.1f} мин")
        if overdue:
            reason_parts.insert(0, "Опоздание > 60 мин")
        if missing_stages:
            reason_parts.append("; ".join(missing_stages))
        delay_reason = " | ".join(reason_parts) if reason_parts else ""

        detailed_orders.append(
            {
                "sale": sale,
                "number": meta.get("number"),
                "restaurant": meta.get("restaurant") or "Не указан",
                "courier": meta.get("courier"),
                "operator": meta.get("operator"),
                "order_type": "Самовывоз" if is_pickup else "Доставка",
                "start_time": start_time.strftime("%Y-%m-%d %H:%M:%S") if start_time else None,
                "done_time": done_time.strftime("%Y-%m-%d %H:%M:%S") if done_time else None,
                "promised_time": promised_time.strftime("%Y-%m-%d %H:%M:%S") if promised_time else None,
                "promised_delta_min": promised_delta_min,
                "total_min": total_min,
                "processing_min": processing_min,
                "cooking_min": cooking_min,
                "assembly_min": assembly_min,
                "delivery_min": None if is_pickup else delivery_min,
                "pickup_wait_min": pickup_wait_min,
                "overdue": overdue,
                "delay_reason": delay_reason,
                "bottleneck_stage": bottleneck_stage,
                "bottleneck_min": bottleneck_min,
            }
        )

    def stage_block(name: str, values: list[float]) -> dict:
        if not values:
            return {"name": name, "count": 0, "avg": None, "p90": None, "max": None}
        return {
            "name": name,
            "count": len(values),
            "avg": statistics.mean(values),
            "p90": _percentile(values, 0.9),
            "max": max(values),
        }

    restaurant_filter = (restaurant_filter or "").strip()
    if restaurant_filter:
        target = norm_status(restaurant_filter)
        detailed_orders = [
            row
            for row in detailed_orders
            if norm_status(row.get("restaurant")) == target
        ]
    detailed_orders = _sort_orders(detailed_orders, sort_mode)

    processing_vals = [float(x["processing_min"]) for x in detailed_orders if isinstance(x.get("processing_min"), (int, float))]
    cooking_vals = [float(x["cooking_min"]) for x in detailed_orders if isinstance(x.get("cooking_min"), (int, float))]
    assembly_vals = [float(x["assembly_min"]) for x in detailed_orders if isinstance(x.get("assembly_min"), (int, float))]
    delivery_vals = [float(x["delivery_min"]) for x in detailed_orders if isinstance(x.get("delivery_min"), (int, float))]
    total_vals = [float(x["total_min"]) for x in detailed_orders if isinstance(x.get("total_min"), (int, float))]

    stages = [
        stage_block("Обработка", processing_vals),
        stage_block("Готовка", cooking_vals),
        stage_block("Сборка", assembly_vals),
        stage_block("Доставка", delivery_vals),
        stage_block("Итого", total_vals),
    ]

    total_orders = len(detailed_orders)
    overdue_orders = [x for x in detailed_orders if x.get("overdue")]
    delivery_orders = [x for x in detailed_orders if x.get("order_type") == "Доставка"]
    pickup_orders = [x for x in detailed_orders if x.get("order_type") == "Самовывоз"]
    delivery_only_vals = [float(x["delivery_min"]) for x in delivery_orders if isinstance(x.get("delivery_min"), (int, float))]

    no_delivery_stage_count = len([x for x in delivery_orders if x.get("delivery_min") is None])
    no_delivery_stage_rate = (
        no_delivery_stage_count / len(delivery_orders) * 100.0
        if delivery_orders
        else 0.0
    )

    filtered_bottlenecks: Counter[str] = Counter(
        str(x.get("bottleneck_stage") or "Не определено") for x in detailed_orders
    )
    bottleneck_rows: list[dict] = []
    for stage, count in filtered_bottlenecks.most_common():
        share = (count / total_orders * 100.0) if total_orders else 0.0
        bottleneck_rows.append({"stage": stage, "count": count, "share": share})

    transitions = [
        {"from": src, "to": dst, "count": cnt}
        for (src, dst), cnt in transitions_counter.most_common(15)
    ]
    phase_stats = []
    for phase_name, values in phase_values_map.items():
        phase_stats.append(stage_block(phase_name, values))

    by_restaurant: dict[str, list[dict]] = defaultdict(list)
    for row in detailed_orders:
        by_restaurant[str(row.get("restaurant") or "Не указан")].append(row)
    hotspots: list[dict] = []
    restaurant_totals: list[dict] = []
    for restaurant, items in by_restaurant.items():
        totals = [float(x["total_min"]) for x in items if isinstance(x.get("total_min"), (int, float))]
        processing = [float(x["processing_min"]) for x in items if isinstance(x.get("processing_min"), (int, float))]
        cooking = [float(x["cooking_min"]) for x in items if isinstance(x.get("cooking_min"), (int, float))]
        assembly = [float(x["assembly_min"]) for x in items if isinstance(x.get("assembly_min"), (int, float))]
        last_mile = []
        for x in items:
            if x.get("order_type") == "Самовывоз":
                if isinstance(x.get("pickup_wait_min"), (int, float)):
                    last_mile.append(float(x["pickup_wait_min"]))
            elif isinstance(x.get("delivery_min"), (int, float)):
                last_mile.append(float(x["delivery_min"]))
        deliveries = [float(x["delivery_min"]) for x in items if isinstance(x.get("delivery_min"), (int, float))]
        overdue_count = sum(1 for x in items if x.get("overdue"))
        restaurant_totals.append(
            {
                "restaurant": restaurant,
                "orders": len(items),
                "overdue_count": overdue_count,
                "overdue_share": (overdue_count / len(items) * 100.0) if items else 0.0,
                "avg_total_min": statistics.mean(totals) if totals else None,
                "p90_total_min": _percentile(totals, 0.9) if totals else None,
                "avg_processing_min": statistics.mean(processing) if processing else None,
                "avg_cooking_min": statistics.mean(cooking) if cooking else None,
                "avg_assembly_min": statistics.mean(assembly) if assembly else None,
                "avg_last_mile_min": statistics.mean(last_mile) if last_mile else None,
            }
        )
        if len(items) < 3:
            continue
        hotspots.append(
            {
                "restaurant": restaurant,
                "orders": len(items),
                "avg_total": statistics.mean(totals) if totals else None,
                "avg_delivery": statistics.mean(deliveries) if deliveries else None,
                "p90_delivery": _percentile(deliveries, 0.9) if deliveries else None,
                "late_share": (overdue_count / len(items) * 100.0) if items else 0.0,
            }
        )
    hotspots.sort(key=lambda x: (-(x["late_share"] or 0.0), -(x["avg_total"] or 0.0), -x["orders"]))
    restaurant_totals.sort(
        key=lambda x: (
            -(x.get("orders") or 0),
            -(x.get("overdue_share") or 0.0),
            str(x.get("restaurant") or ""),
        )
    )

    problem_orders = overdue_orders

    load_map: dict[str, dict] = defaultdict(lambda: {"hour": "", "count": 0, "overdue_count": 0, "total_minutes": 0.0})
    for row in detailed_orders:
        start = parse_dt(row.get("start_time"))
        if start is None:
            continue
        hour_key = f"{start.hour:02d}:00"
        bucket = load_map[hour_key]
        bucket["hour"] = hour_key
        bucket["count"] += 1
        if row.get("overdue"):
            bucket["overdue_count"] += 1
        if isinstance(row.get("total_min"), (int, float)):
            bucket["total_minutes"] += float(row["total_min"])
    load_by_hour = []
    for hour in sorted(load_map.keys()):
        b = load_map[hour]
        avg_total = (b["total_minutes"] / b["count"]) if b["count"] else None
        load_by_hour.append(
            {
                "hour": hour,
                "count": b["count"],
                "overdue_count": b["overdue_count"],
                "avg_total_min": avg_total,
                "total_minutes": b["total_minutes"],
            }
        )

    return {
        "kpi": {
            "orders": total_orders,
            "avg_total_min": statistics.mean(total_vals) if total_vals else None,
            "p90_total_min": _percentile(total_vals, 0.9) if total_vals else None,
            "avg_delivery_min": statistics.mean(delivery_only_vals) if delivery_only_vals else None,
            "p90_delivery_min": _percentile(delivery_only_vals, 0.9) if delivery_only_vals else None,
            "overdue_count": len(overdue_orders),
            "overdue_rate": (len(overdue_orders) / total_orders * 100.0) if total_orders else 0.0,
            "delivery_orders": len(delivery_orders),
            "pickup_orders": len(pickup_orders),
            "orders_with_statuses": len(events_by_sale),
            "no_delivery_stage_count": no_delivery_stage_count,
            "no_delivery_stage_rate": no_delivery_stage_rate,
        },
        "thresholds": {
            "overdue_total_min": overdue_total_threshold,
            "late_delivery_min": late_delivery_threshold,
            "critical_delivery_min": critical_delivery_threshold,
        },
        "stages": stages,
        "bottlenecks": bottleneck_rows,
        "hotspots": hotspots[:20],
        "restaurant_totals": restaurant_totals,
        "problem_orders": problem_orders,
        "status_flow": {
            "orders_with_statuses": len(events_by_sale),
            "no_delivery_stage_count": no_delivery_stage_count,
            "no_delivery_stage_rate": no_delivery_stage_rate,
            "transitions": transitions,
            "phase_stats": phase_stats,
        },
        "load_by_hour": load_by_hour,
        "orders": detailed_orders,
        "restaurant_filter": restaurant_filter or None,
        "sort_mode": (sort_mode or "restaurant_asc"),
    }


PDF_FONT_NAME: str | None = None


def _ensure_pdf_font() -> str:
    global PDF_FONT_NAME
    if PDF_FONT_NAME:
        return PDF_FONT_NAME
    candidates = [
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
    ]
    for font_path in candidates:
        p = Path(font_path)
        if p.exists():
            try:
                pdfmetrics.registerFont(TTFont("UIFont", str(p)))
                PDF_FONT_NAME = "UIFont"
                return PDF_FONT_NAME
            except Exception:
                continue
    PDF_FONT_NAME = "Helvetica"
    return PDF_FONT_NAME


def _list_restaurants(file_path: Path) -> list[str]:
    wb = load_workbook(filename=str(file_path), data_only=True, read_only=True)
    if "Реестр" not in wb.sheetnames:
        return []
    rows = _sheet_dict_rows(wb["Реестр"])
    names = set()
    for row in rows:
        name = (
            row.get("WarehouseName")
            or row.get("CompanyName")
            or row.get("RealCompanyName")
            or row.get("OriginCompanyName")
            or ""
        )
        text = str(name).strip()
        if text:
            names.add(text)
    return sorted(names)


def _build_pdf_report(file_path: Path, restaurant_filter: str | None, sort_mode: str | None) -> bytes:
    payload = build_analytics_payload(file_path, restaurant_filter=restaurant_filter, sort_mode=sort_mode)
    rows = payload.get("orders") or []
    restaurant_totals = payload.get("restaurant_totals") or []
    kpi = payload.get("kpi") or {}
    threshold = float((payload.get("thresholds") or {}).get("overdue_total_min", 60.0))

    font_name = _ensure_pdf_font()
    styles = getSampleStyleSheet()
    for style_key in ("Normal", "Title", "Heading2"):
        styles[style_key].fontName = font_name

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
        title="Отчет по доставке",
    )
    story = []
    date_part = file_path.stem.replace("order_status_history_", "")
    restaurant_part = restaurant_filter or "Все рестораны"
    story.append(Paragraph(f"Отчет доставки за {date_part}", styles["Title"]))
    story.append(Paragraph(f"Ресторан: {restaurant_part}", styles["Normal"]))
    story.append(
        Paragraph(
            (
                f"Заказов: {kpi.get('orders', 0)} | "
                f"Опозданий (>{int(threshold)} мин): {kpi.get('overdue_count', 0)} ({(kpi.get('overdue_rate') or 0):.1f}%) | "
                f"Среднее время заказа: {(kpi.get('avg_total_min') or 0):.1f} мин"
            ),
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 8))

    if restaurant_totals:
        summary_header = ["Ресторан", "Заказы", "Опозд.", "Доля %", "Avg итого", "P90 итого"]
        summary_rows = [summary_header]
        for row in restaurant_totals[:25]:
            summary_rows.append(
                [
                    str(row.get("restaurant") or "—"),
                    str(int(row.get("orders") or 0)),
                    str(int(row.get("overdue_count") or 0)),
                    f"{float(row.get('overdue_share') or 0):.1f}",
                    "—" if row.get("avg_total_min") is None else f"{float(row.get('avg_total_min')):.1f}",
                    "—" if row.get("p90_total_min") is None else f"{float(row.get('p90_total_min')):.1f}",
                ]
            )
        story.append(Paragraph("Итоги по ресторанам", styles["Heading2"]))
        rest_table = Table(summary_rows, colWidths=[200, 55, 55, 55, 70, 70], repeatRows=1)
        rest_style = TableStyle(
            [
                ("FONT", (0, 0), (-1, -1), font_name, 8),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#c7cdd4")),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ]
        )
        rest_table.setStyle(rest_style)
        story.append(rest_table)
        story.append(Spacer(1, 8))

    header = [
        "Ресторан",
        "Заказ",
        "План прибытия",
        "Факт (выполнен)",
        "Δ план/факт (мин)",
        "Итого",
        "Обработка",
        "Готовка",
        "Сборка",
        "Доставка/Выдача",
        "Причина",
    ]
    table_rows = [header]
    for row in rows:
        last_stage = row.get("pickup_wait_min") if row.get("order_type") == "Самовывоз" else row.get("delivery_min")
        delta = row.get("promised_delta_min")
        delta_text = "—"
        if isinstance(delta, (int, float)):
            delta_text = f"{delta:+.1f}"
        table_rows.append(
            [
                str(row.get("restaurant") or "—"),
                str(row.get("number") or row.get("sale") or "—"),
                str(row.get("promised_time") or "—"),
                str(row.get("done_time") or "—"),
                delta_text,
                "—" if row.get("total_min") is None else f"{float(row.get('total_min')):.1f}",
                "—" if row.get("processing_min") is None else f"{float(row.get('processing_min')):.1f}",
                "—" if row.get("cooking_min") is None else f"{float(row.get('cooking_min')):.1f}",
                "—" if row.get("assembly_min") is None else f"{float(row.get('assembly_min')):.1f}",
                "—" if last_stage is None else f"{float(last_stage):.1f}",
                str(row.get("delay_reason") or "—"),
            ]
        )

    widths = [130, 66, 88, 88, 72, 45, 50, 50, 50, 70, 170]
    table = Table(table_rows, colWidths=widths, repeatRows=1)
    style = TableStyle(
        [
            ("FONT", (0, 0), (-1, -1), font_name, 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#c7cdd4")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (4, 1), (9, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
        ]
    )
    for idx, row in enumerate(rows, start=1):
        is_overdue = bool(isinstance(row.get("total_min"), (int, float)) and float(row.get("total_min")) > threshold)
        is_late_plan = bool(isinstance(row.get("promised_delta_min"), (int, float)) and float(row.get("promised_delta_min")) > 0)
        if is_overdue or is_late_plan:
            style.add("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#fff1f2"))
            style.add("TEXTCOLOR", (4, idx), (5, idx), colors.HexColor("#b91c1c"))
    table.setStyle(style)
    story.append(table)
    doc.build(story)
    return buffer.getvalue()


def _build_excel_report(file_path: Path, restaurant_filter: str | None, sort_mode: str | None) -> bytes:
    payload = build_analytics_payload(file_path, restaurant_filter=restaurant_filter, sort_mode=sort_mode)
    rows = payload.get("orders") or []
    restaurant_totals = payload.get("restaurant_totals") or []
    kpi = payload.get("kpi") or {}
    threshold = float((payload.get("thresholds") or {}).get("overdue_total_min", 60.0))

    wb = Workbook()
    ws_kpi = wb.active
    ws_kpi.title = "Сводка"
    ws_kpi.append(["Показатель", "Значение"])
    ws_kpi.append(["Дата", file_path.stem.replace("order_status_history_", "")])
    ws_kpi.append(["Ресторан", restaurant_filter or "Все рестораны"])
    ws_kpi.append(["Сортировка", sort_mode or "restaurant_asc"])
    ws_kpi.append(["Заказов", kpi.get("orders")])
    ws_kpi.append(["Опозданий > 60 мин", kpi.get("overdue_count")])
    ws_kpi.append(["Доля опозданий, %", kpi.get("overdue_rate")])
    ws_kpi.append(["Среднее время заказа, мин", kpi.get("avg_total_min")])
    ws_kpi.append(["P90 времени заказа, мин", kpi.get("p90_total_min")])
    ws_kpi.append(["Средняя доставка, мин", kpi.get("avg_delivery_min")])
    ws_kpi.append(["P90 доставка, мин", kpi.get("p90_delivery_min")])

    ws_orders = wb.create_sheet("Заказы")
    ws_orders.append(
        [
            "Ресторан",
            "Заказ",
            "Тип",
            "Старт",
            "К какому времени",
            "Факт (выполнен)",
            "Δ план/факт (мин)",
            "Итого (мин)",
            "Обработка",
            "Готовка",
            "Сборка",
            "Доставка/Выдача",
            "Курьер",
            "Оператор",
            "Узкое место",
            "Причина",
            "Опоздание > 60 мин",
        ]
    )
    for row in rows:
        last_stage = row.get("pickup_wait_min") if row.get("order_type") == "Самовывоз" else row.get("delivery_min")
        ws_orders.append(
            [
                row.get("restaurant"),
                row.get("number") or row.get("sale"),
                row.get("order_type"),
                row.get("start_time"),
                row.get("promised_time"),
                row.get("done_time"),
                row.get("promised_delta_min"),
                row.get("total_min"),
                row.get("processing_min"),
                row.get("cooking_min"),
                row.get("assembly_min"),
                last_stage,
                row.get("courier"),
                row.get("operator"),
                row.get("bottleneck_stage"),
                row.get("delay_reason"),
                "ДА" if row.get("total_min") is not None and float(row.get("total_min")) > threshold else "",
            ]
        )

    ws_problem = wb.create_sheet("Проблемные")
    ws_problem.append(
        [
            "Ресторан",
            "Заказ",
            "Тип",
            "К какому времени",
            "Факт (выполнен)",
            "Δ план/факт (мин)",
            "Итого (мин)",
            "Этапы (обраб/готов/сбор/дост)",
            "Узкое место",
            "Причина",
        ]
    )
    for row in rows:
        total = row.get("total_min")
        if not isinstance(total, (int, float)) or float(total) <= threshold:
            continue
        last_stage = row.get("pickup_wait_min") if row.get("order_type") == "Самовывоз" else row.get("delivery_min")
        ws_problem.append(
            [
                row.get("restaurant"),
                row.get("number") or row.get("sale"),
                row.get("order_type"),
                row.get("promised_time"),
                row.get("done_time"),
                row.get("promised_delta_min"),
                total,
                (
                    f"{(row.get('processing_min') or 0):.1f} / "
                    f"{(row.get('cooking_min') or 0):.1f} / "
                    f"{(row.get('assembly_min') or 0):.1f} / "
                    f"{(last_stage or 0):.1f}"
                ),
                f"{row.get('bottleneck_stage') or '—'} ({(row.get('bottleneck_min') or 0):.1f})",
                row.get("delay_reason"),
            ]
        )

    ws_rest = wb.create_sheet("Итоги по ресторанам")
    ws_rest.append(
        [
            "Ресторан",
            "Заказы",
            "Опозданий",
            "Доля опозданий, %",
            "Avg итого, мин",
            "P90 итого, мин",
            "Avg обработка, мин",
            "Avg готовка, мин",
            "Avg сборка, мин",
            "Avg доставка/выдача, мин",
        ]
    )
    for row in restaurant_totals:
        ws_rest.append(
            [
                row.get("restaurant"),
                row.get("orders"),
                row.get("overdue_count"),
                row.get("overdue_share"),
                row.get("avg_total_min"),
                row.get("p90_total_min"),
                row.get("avg_processing_min"),
                row.get("avg_cooking_min"),
                row.get("avg_assembly_min"),
                row.get("avg_last_mile_min"),
            ]
        )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt: str, *args) -> None:
        return

    def send_json(self, data: dict, status: int = 200) -> None:
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_text(self, text: str, status: int = 200) -> None:
        body = text.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self) -> None:  # noqa: N802
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")
        self.end_headers()

    def serve_static(self, rel_path: str) -> None:
        file_path = (WEB_DIR / rel_path).resolve()
        if not str(file_path).startswith(str(WEB_DIR.resolve())):
            self.send_text("forbidden", status=403)
            return
        if not file_path.exists() or not file_path.is_file():
            self.send_text("not found", status=404)
            return

        content = file_path.read_bytes()
        ctype = "text/plain; charset=utf-8"
        if file_path.suffix == ".html":
            ctype = "text/html; charset=utf-8"
        elif file_path.suffix == ".css":
            ctype = "text/css; charset=utf-8"
        elif file_path.suffix == ".js":
            ctype = "application/javascript; charset=utf-8"

        self.send_response(200)
        self.send_header("Content-Type", ctype)
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def do_GET(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        path = parsed.path

        if path == "/api/health":
            self.send_json({"ok": True, "time": now_iso()})
            return

        if path == "/api/latest":
            with JOB_LOCK:
                job = JOBS.get(LATEST_JOB_ID) if LATEST_JOB_ID else None
                if not job:
                    self.send_json({"job": None})
                    return
                self.send_json({"job": {k: v for k, v in job.items() if k != "logs"}})
            return

        if path.startswith("/api/job/"):
            job_id = path.split("/api/job/", 1)[1]
            from_idx = 0
            qs = parse_qs(parsed.query)
            if "from" in qs:
                try:
                    from_idx = max(0, int(qs["from"][0]))
                except ValueError:
                    from_idx = 0

            with JOB_LOCK:
                job = JOBS.get(job_id)
                if not job:
                    self.send_json({"error": "job not found"}, status=404)
                    return
                logs = job.get("logs", [])
                payload = {
                    "id": job["id"],
                    "date": job["date"],
                    "status": job["status"],
                    "started_at": job["started_at"],
                    "ended_at": job.get("ended_at"),
                    "output_path": job.get("output_path"),
                    "error": job.get("error"),
                    "logs": logs[from_idx:],
                    "log_size": len(logs),
                }
            self.send_json(payload)
            return

        if path == "/api/download":
            qs = parse_qs(parsed.query)
            requested_id = (qs.get("job_id", [None])[0] or None)
            if requested_id is not None:
                requested_id = str(requested_id).strip() or None
            requested_date = (qs.get("date", [None])[0] or None)
            restaurant_filter = (qs.get("restaurant", [None])[0] or None)
            sort_mode = (qs.get("sort", [None])[0] or None)
            if requested_date is not None:
                requested_date = str(requested_date).strip() or None
            if restaurant_filter is not None:
                restaurant_filter = str(restaurant_filter).strip() or None
            if sort_mode is not None:
                sort_mode = str(sort_mode).strip() or None

            try:
                if requested_date:
                    file_path = _resolve_output_by_date(requested_date)
                else:
                    _, _, file_path = _resolve_job_output(requested_id)
                content = _build_excel_report(
                    file_path,
                    restaurant_filter=restaurant_filter,
                    sort_mode=sort_mode,
                )
            except RuntimeError as err:
                message = str(err)
                if message == "job not found":
                    self.send_json({"error": message}, status=404)
                elif message == "job is not completed successfully":
                    self.send_json({"error": message}, status=409)
                elif message == "date must be YYYY-MM-DD":
                    self.send_json({"error": message}, status=400)
                elif message == "forbidden output path":
                    self.send_json({"error": message}, status=403)
                else:
                    self.send_json({"error": message}, status=404)
                return
            except Exception as err:  # noqa: BLE001
                self.send_json({"error": f"excel report failed: {err}"}, status=500)
                return

            suffix = requested_date or "latest"
            if restaurant_filter:
                safe_name = re.sub(r"[^a-zA-Z0-9а-яА-ЯёЁ_-]+", "_", restaurant_filter)[:40]
                suffix = f"{suffix}_{safe_name}"
            filename = f"delivery_report_{suffix}.xlsx"
            self.send_response(200)
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(len(content)))
            self.end_headers()
            self.wfile.write(content)
            return

        if path == "/api/restaurants":
            qs = parse_qs(parsed.query)
            requested_date = (qs.get("date", [None])[0] or None)
            if requested_date is not None:
                requested_date = str(requested_date).strip() or None
            if not requested_date:
                self.send_json({"error": "date is required"}, status=400)
                return
            try:
                file_path = _resolve_output_by_date(requested_date)
                restaurants = _list_restaurants(file_path)
            except RuntimeError as err:
                message = str(err)
                code = 404
                if message == "date must be YYYY-MM-DD":
                    code = 400
                elif message == "forbidden output path":
                    code = 403
                self.send_json({"error": message}, status=code)
                return
            self.send_json({"date": requested_date, "restaurants": restaurants})
            return

        if path == "/api/report_pdf":
            qs = parse_qs(parsed.query)
            requested_date = (qs.get("date", [None])[0] or None)
            restaurant_filter = (qs.get("restaurant", [None])[0] or None)
            sort_mode = (qs.get("sort", [None])[0] or None)
            if requested_date is not None:
                requested_date = str(requested_date).strip() or None
            if restaurant_filter is not None:
                restaurant_filter = str(restaurant_filter).strip() or None
            if sort_mode is not None:
                sort_mode = str(sort_mode).strip() or None
            if not requested_date:
                self.send_json({"error": "date is required"}, status=400)
                return
            try:
                file_path = _resolve_output_by_date(requested_date)
                content = _build_pdf_report(
                    file_path,
                    restaurant_filter=restaurant_filter,
                    sort_mode=sort_mode,
                )
            except RuntimeError as err:
                message = str(err)
                code = 404
                if message == "date must be YYYY-MM-DD":
                    code = 400
                elif message == "forbidden output path":
                    code = 403
                self.send_json({"error": message}, status=code)
                return
            except Exception as err:  # noqa: BLE001
                self.send_json({"error": f"pdf report failed: {err}"}, status=500)
                return

            suffix = requested_date
            if restaurant_filter:
                safe_name = re.sub(r"[^a-zA-Z0-9а-яА-ЯёЁ_-]+", "_", restaurant_filter)[:40]
                suffix = f"{suffix}_{safe_name}"
            filename = f"delivery_report_{suffix}.pdf"
            self.send_response(200)
            self.send_header("Content-Type", "application/pdf")
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(len(content)))
            self.end_headers()
            self.wfile.write(content)
            return

        if path == "/api/analytics":
            qs = parse_qs(parsed.query)
            requested_id = (qs.get("job_id", [None])[0] or None)
            if requested_id is not None:
                requested_id = str(requested_id).strip() or None
            requested_date = (qs.get("date", [None])[0] or None)
            if requested_date is not None:
                requested_date = str(requested_date).strip() or None
            restaurant_filter = (qs.get("restaurant", [None])[0] or None)
            sort_mode = (qs.get("sort", [None])[0] or None)
            if restaurant_filter is not None:
                restaurant_filter = str(restaurant_filter).strip() or None
            if sort_mode is not None:
                sort_mode = str(sort_mode).strip() or None
            try:
                if requested_date:
                    file_path = _resolve_output_by_date(requested_date)
                    target_id = None
                    job = {"date": requested_date}
                else:
                    target_id, job, file_path = _resolve_job_output(requested_id)
                payload = build_analytics_payload(
                    file_path,
                    restaurant_filter=restaurant_filter,
                    sort_mode=sort_mode,
                )
            except RuntimeError as err:
                message = str(err)
                if message == "job not found":
                    self.send_json({"error": message}, status=404)
                elif message == "job is not completed successfully":
                    self.send_json({"error": message}, status=409)
                elif message == "date must be YYYY-MM-DD":
                    self.send_json({"error": message}, status=400)
                elif message == "forbidden output path":
                    self.send_json({"error": message}, status=403)
                else:
                    self.send_json({"error": message}, status=404)
                return
            except Exception as err:  # noqa: BLE001
                self.send_json({"error": f"analytics parse failed: {err}"}, status=500)
                return

            payload.update(
                {
                    "job_id": target_id,
                    "date": job.get("date"),
                    "output_path": str(file_path),
                    "generated_at": now_iso(),
                    "restaurant_filter": restaurant_filter,
                }
            )
            self.send_json(payload)
            return

        if path == "/":
            self.serve_static("index.html")
            return

        if path in {"/styles.css", "/app.js", "/version.js"}:
            self.serve_static(path.lstrip("/"))
            return

        self.send_text("not found", status=404)

    def do_POST(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        if parsed.path == "/api/stop":
            content_len = int(self.headers.get("Content-Length", "0"))
            raw = self.rfile.read(content_len) if content_len else b"{}"
            try:
                payload = json.loads(raw.decode("utf-8"))
            except json.JSONDecodeError:
                payload = {}
            job_id = payload.get("job_id")
            if job_id is not None:
                job_id = str(job_id).strip() or None
            try:
                result = stop_job(job_id)
            except RuntimeError as err:
                self.send_json({"error": str(err)}, status=409)
                return
            self.send_json(result, status=HTTPStatus.ACCEPTED)
            return

        if parsed.path != "/api/run":
            self.send_text("not found", status=404)
            return

        content_len = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(content_len) if content_len else b"{}"

        try:
            payload = json.loads(raw.decode("utf-8"))
        except json.JSONDecodeError:
            self.send_json({"error": "invalid json"}, status=400)
            return

        date_value = str(payload.get("date", "")).strip()
        try:
            datetime.strptime(date_value, "%Y-%m-%d")
        except ValueError:
            self.send_json({"error": "date must be YYYY-MM-DD"}, status=400)
            return

        try:
            job = create_job(date_value)
        except RuntimeError as err:
            self.send_json({"error": str(err)}, status=409)
            return

        self.send_json({"job_id": job["id"], "status": job["status"]}, status=HTTPStatus.ACCEPTED)


def main() -> int:
    if "--worker" in sys.argv:
        try:
            idx = sys.argv.index("--date")
            date_value = sys.argv[idx + 1]
        except Exception:
            print("worker mode requires --date YYYY-MM-DD")
            return 2
        return run_embedded_worker(str(date_value))

    worker_exe = Path(sys.executable).resolve().with_name("export_delivery_statuses.exe")
    if not WEB_DIR.exists():
        print(f"web dir not found: {WEB_DIR}")
        return 2
    if not EXPORT_SCRIPT.exists() and not worker_exe.exists() and not getattr(sys, "frozen", False):
        print(f"export script not found: {EXPORT_SCRIPT}")
        print(f"worker exe not found: {worker_exe}")
        return 2

    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"Web UI: http://{HOST}:{PORT}")
    print("Press Ctrl+C to stop")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
