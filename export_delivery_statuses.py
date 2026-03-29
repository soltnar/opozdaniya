#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import copy
import json
import os
import re
import sys
import time
import math
import statistics
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from playwright.sync_api import BrowserContext, Page, sync_playwright

DELIVERY_URL = "https://rest.saby.ru/page/delivery"
DEFAULT_HAR_CANDIDATES = [
    "/Users/macbook/Downloads/rest.saby.ru_dost.har",
    "/Users/macbook/Downloads/rest.saby.ru history.har",
]

DEFAULT_SERVICE_URL = "https://rest.saby.ru/service/"

ALLOWED_BASE_HEADERS = {
    "accept",
    "content-type",
    "origin",
    "referer",
    "x-adaptive",
    "x-osversion",
    "x-requested-with",
    "x-saby-appid",
    "x-saby-appversion",
    "x-saby-cfgid",
}

DEFAULT_BASE_HEADERS = {
    "accept": "application/json, text/javascript, */*; q=0.01",
    "content-type": "application/json; charset=UTF-8",
    "origin": "https://rest.saby.ru",
    "referer": DELIVERY_URL,
    "x-adaptive": "false",
    "x-requested-with": "XMLHttpRequest",
}

STATUS_CHANGE_PATTERN = re.compile(
    r'Изменен статус заказа:\s*["«](?P<from>.*?)["»]\s*->\s*["«](?P<to>.*?)["»]'
)


@dataclass
class TemplateBundle:
    service_url: str
    base_headers: dict[str, str]
    sale_payload_template: dict[str, Any]
    sale_called_method: str
    sale_position_field: str
    sale_position_type: str
    har_sale_delivery_context: bool
    history_payload_template: dict[str, Any]
    history_called_method: str
    history_position_field: str
    history_position_type: str


class SabyRpcClient:
    def __init__(self, context: BrowserContext, service_url: str, base_headers: dict[str, str]) -> None:
        self._context = context
        self._service_url = service_url
        self._base_headers = base_headers

    def call(self, payload: dict[str, Any], called_method: str) -> dict[str, Any]:
        headers = dict(self._base_headers)
        body_method = payload.get("method", "")
        headers["x-calledmethod"] = called_method
        headers["x-originalmethodname"] = base64.b64encode(body_method.encode("utf-8")).decode("ascii")
        last_err: Exception | None = None
        for attempt in range(1, 4):
            try:
                response = self._context.request.post(
                    self._service_url,
                    data=json.dumps(payload, ensure_ascii=False),
                    headers=headers,
                    timeout=60_000,
                )
                if response.status >= 500:
                    raise RuntimeError(
                        f"RPC HTTP error {response.status}: {response.text()[:500]}"
                    )
                if response.status >= 400:
                    raise RuntimeError(
                        f"RPC HTTP error {response.status}: {response.text()[:500]}"
                    )

                data = response.json()
                if "error" in data:
                    err = data["error"]
                    raise RuntimeError(f"RPC error: {json.dumps(err, ensure_ascii=False)}")

                result = data.get("result")
                if not isinstance(result, dict):
                    raise RuntimeError(f"RPC result has unexpected format: {type(result).__name__}")
                return result
            except Exception as err:  # noqa: BLE001
                last_err = err
                message = str(err)
                transient = (
                    "RPC HTTP error 500" in message
                    or "RPC HTTP error 502" in message
                    or "RPC HTTP error 503" in message
                    or "RPC HTTP error 504" in message
                )
                if attempt < 3 and transient:
                    time.sleep(0.8 * attempt)
                    continue
                raise
        if last_err is not None:
            raise last_err
        raise RuntimeError("RPC call failed")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Выгрузка истории статусов заказов доставки из Saby в Excel"
    )
    parser.add_argument(
        "--date",
        required=True,
        help="Дата выборки в формате YYYY-MM-DD",
    )
    parser.add_argument(
        "--har",
        default="auto",
        help="Путь к HAR-файлу с примерами запросов (или auto для авто-выбора)",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Путь к выходному xlsx. По умолчанию exports/order_status_history_<date>.xlsx",
    )
    parser.add_argument(
        "--profile-dir",
        default=".saby_profile",
        help="Каталог профиля Chromium для сохранения сессии",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Запустить браузер в headless-режиме",
    )
    parser.add_argument(
        "--order-page-limit",
        type=int,
        default=25,
        help="Размер страницы при загрузке реестра заказов",
    )
    parser.add_argument(
        "--history-page-limit",
        type=int,
        default=24,
        help="Размер страницы при загрузке истории заказа",
    )
    parser.add_argument(
        "--max-orders",
        type=int,
        default=None,
        help="Ограничить число заказов (для отладки)",
    )
    parser.add_argument(
        "--max-order-pages",
        type=int,
        default=None,
        help="Ограничить число страниц реестра заказов (для отладки)",
    )
    parser.add_argument(
        "--max-history-pages",
        type=int,
        default=None,
        help="Ограничить число страниц истории на один заказ (для отладки)",
    )
    parser.add_argument(
        "--allow-heuristic-fallback",
        action="store_true",
        help="Разрешить эвристический подбор состава заказов, если runtime ProductStateId=Done не пойман",
    )
    parser.add_argument(
        "--align-to-ui-count",
        action="store_true",
        help="Включить подгонку состава заказов под UI-счетчик 'Выполнен/Отмечено' (по умолчанию отключено)",
    )
    return parser.parse_args()


def resolve_har_path(raw_value: str) -> Path | None:
    value = (raw_value or "").strip()
    if value and value.lower() != "auto":
        candidate = Path(value).expanduser()
        return candidate if candidate.exists() else None
    for item in DEFAULT_HAR_CANDIDATES:
        candidate = Path(item).expanduser()
        if candidate.exists():
            return candidate
    return None


def make_default_history_payload_template() -> dict[str, Any]:
    return {
        "jsonrpc": "2.0",
        "protocol": 7,
        "id": 1,
        "method": "История.History_Of_Instance",
        "params": {
            "Фильтр": {
                "d": [None, None, ["Изменение статуса заказа"]],
                "s": [
                    {"t": "Строка", "n": "GUID"},
                    {"t": "Число целое", "n": "ИдО"},
                    {"t": "Массив", "n": "Действие"},
                ],
                "_type": "record",
                "f": 0,
            },
            "Навигация": {
                "d": ["forward", True, 24, None],
                "s": [
                    {"t": "Строка", "n": "Direction"},
                    {"t": "Логическое", "n": "HasMore"},
                    {"t": "Число целое", "n": "Limit"},
                    {"t": "Запись", "n": "Position"},
                ],
                "_type": "record",
                "f": 0,
            },
        },
    }


def make_default_sale_payload_template() -> dict[str, Any]:
    return {
        "jsonrpc": "2.0",
        "protocol": 7,
        "id": 1,
        "method": "SaleOrder.List",
        "params": {
            "Фильтр": {
                "d": [
                    "2000-01-01 00:00:00",
                    "2000-01-01 23:59:59",
                    "Done",
                    2,
                ],
                "s": [
                    {"t": "Строка", "n": "DateTimeStartWTZ"},
                    {"t": "Строка", "n": "DateTimeEndWTZ"},
                    {"t": "Строка", "n": "ProductStateId"},
                    {"t": "Число целое", "n": "Reglament"},
                ],
                "_type": "record",
                "f": 0,
            },
            "Навигация": {
                "d": ["bothways", True, 25, None],
                "s": [
                    {"t": "Строка", "n": "Direction"},
                    {"t": "Логическое", "n": "HasMore"},
                    {"t": "Число целое", "n": "Limit"},
                    {"t": "Запись", "n": "Position"},
                ],
                "_type": "record",
                "f": 0,
            },
            "Сортировка": {
                "d": [[False, "NextDateWTZText", True]],
                "s": [
                    {"t": "Логическое", "n": "l"},
                    {"t": "Строка", "n": "n"},
                    {"t": "Логическое", "n": "o"},
                ],
                "_type": "recordset",
                "f": 0,
            },
        },
    }


def make_runtime_fallback_templates(
    runtime_sale_payload: dict[str, Any] | None,
    runtime_sale_called_method: str | None,
) -> TemplateBundle:
    sale_payload: dict[str, Any]
    sale_called_method = runtime_sale_called_method or "SaleOrder.List"
    if isinstance(runtime_sale_payload, dict) and isinstance(runtime_sale_payload.get("params"), dict):
        sale_payload = copy.deepcopy(runtime_sale_payload)
        sale_payload.setdefault("method", "SaleOrder.List")
    else:
        sale_payload = make_default_sale_payload_template()

    history_payload = make_default_history_payload_template()

    sale_position_field, sale_position_type = extract_position_meta(
        sale_payload,
        default_field_name="NextDateWTZText",
        default_field_type="Строка",
    )
    history_position_field, history_position_type = extract_position_meta(
        history_payload,
        default_field_name="_time",
        default_field_type="Дата и время",
    )

    return TemplateBundle(
        service_url=DEFAULT_SERVICE_URL,
        base_headers=dict(DEFAULT_BASE_HEADERS),
        sale_payload_template=sale_payload,
        sale_called_method=sale_called_method,
        sale_position_field=sale_position_field,
        sale_position_type=sale_position_type,
        har_sale_delivery_context=False,
        history_payload_template=history_payload,
        history_called_method="Istoriya.History_Of_Instance",
        history_position_field=history_position_field,
        history_position_type=history_position_type,
    )


def parse_date(value: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError("Дата должна быть в формате YYYY-MM-DD") from exc


def read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def get_header(headers: list[dict[str, Any]], name: str) -> str | None:
    name_lower = name.lower()
    for item in headers:
        if str(item.get("name", "")).lower() == name_lower:
            return str(item.get("value", ""))
    return None


def headers_to_base(headers: list[dict[str, Any]]) -> dict[str, str]:
    out = dict(DEFAULT_BASE_HEADERS)
    for item in headers:
        key = str(item.get("name", "")).lower()
        if key in ALLOWED_BASE_HEADERS:
            out[key] = str(item.get("value", ""))
    return out


def record_field_names(record: dict[str, Any]) -> list[str]:
    result: list[str] = []
    for raw in record.get("s", []):
        if isinstance(raw, dict):
            result.append(str(raw.get("n", "")))
        else:
            result.append(str(raw))
    return result


def record_field_index(record: dict[str, Any], field_name: str) -> int | None:
    names = record_field_names(record)
    try:
        return names.index(field_name)
    except ValueError:
        return None


def set_record_field(record: dict[str, Any], field_name: str, value: Any) -> bool:
    index = record_field_index(record, field_name)
    if index is None:
        return False
    values = record.setdefault("d", [])
    while len(values) <= index:
        values.append(None)
    values[index] = value
    return True


def remove_record_field(record: dict[str, Any], field_name: str) -> bool:
    index = record_field_index(record, field_name)
    if index is None:
        return False
    specs = record.get("s", [])
    values = record.get("d", [])
    if isinstance(specs, list) and index < len(specs):
        del specs[index]
    if isinstance(values, list) and index < len(values):
        del values[index]
    return True


def record_to_map(record: dict[str, Any] | None) -> dict[str, Any]:
    if not isinstance(record, dict):
        return {}
    names = record_field_names(record)
    values = record.get("d", [])
    out: dict[str, Any] = {}
    for idx, name in enumerate(names):
        out[name] = values[idx] if idx < len(values) else None
    return out


def copy_record_field(dst_record: dict[str, Any], src_record: dict[str, Any], field_name: str) -> bool:
    src_idx = record_field_index(src_record, field_name)
    if src_idx is None:
        return False
    src_values = src_record.get("d", [])
    src_value = src_values[src_idx] if src_idx < len(src_values) else None
    return set_record_field(dst_record, field_name, copy.deepcopy(src_value))


def merge_runtime_filter_context(
    template_payload: dict[str, Any],
    runtime_payload: dict[str, Any],
) -> int:
    template_filter = template_payload.get("params", {}).get("Фильтр")
    runtime_filter = runtime_payload.get("params", {}).get("Фильтр")
    if not isinstance(template_filter, dict) or not isinstance(runtime_filter, dict):
        return 0

    fields_to_copy = (
        "CRMFilter",
        "Company",
        "OurOrgFilter",
        "Sales",
        "SalesPoints",
        "Workplaces",
        "Warehouses",
        "ProductStateFields",
        "NotFields",
        "NameLength",
        "ShowMiniPaymentsJSON",
        "ShowSaleNomenclature",
    )

    applied = 0
    for field_name in fields_to_copy:
        if copy_record_field(template_filter, runtime_filter, field_name):
            applied += 1
    return applied


def get_record_field(record: dict[str, Any], field_name: str) -> Any:
    index = record_field_index(record, field_name)
    if index is None:
        return None
    values = record.get("d", [])
    if index >= len(values):
        return None
    return values[index]


def pick_sale_template(
    entries: list[dict[str, Any]]
) -> tuple[dict[str, Any], dict[str, Any], str, str, str, bool]:
    candidates: list[tuple[int, dict[str, Any], dict[str, Any], str, str, str, bool]] = []

    for entry in entries:
        req = entry.get("request", {})
        called_method = get_header(req.get("headers", []), "x-calledmethod")
        referer = get_header(req.get("headers", []), "referer")
        if called_method != "SaleOrder.List":
            continue

        post_text = req.get("postData", {}).get("text")
        if not isinstance(post_text, str) or not post_text:
            continue

        try:
            payload = json.loads(post_text)
        except json.JSONDecodeError:
            continue

        params = payload.get("params")
        if not isinstance(params, dict):
            continue

        filter_record = params.get("Фильтр")
        if not isinstance(filter_record, dict):
            continue

        names = record_field_names(filter_record)
        idx_state = record_field_index(filter_record, "ProductStateId")
        if idx_state is None:
            continue

        state_val = get_record_field(filter_record, "ProductStateId")
        if state_val != "Done":
            continue

        score = 0
        if "Sales" not in names:
            score += 5
        if isinstance(params.get("Навигация"), dict):
            score += 3
        if params.get("Сортировка") is not None:
            score += 2

        nav = params.get("Навигация")
        if isinstance(nav, dict):
            nav_values = nav.get("d", [])
            if isinstance(nav_values, list) and len(nav_values) > 3 and isinstance(nav_values[3], dict):
                score += 1

        is_delivery_context = "/page/delivery" in referer
        if is_delivery_context:
            score += 20
        elif "/page/order-list" in referer:
            score -= 5

        candidates.append(
            (
                score,
                entry,
                payload,
                called_method,
                payload.get("method", "SaleOrder.List"),
                req.get("url", ""),
                is_delivery_context,
            )
        )

    if not candidates:
        raise RuntimeError("Не найден подходящий шаблон SaleOrder.List в HAR")

    candidates.sort(key=lambda item: item[0], reverse=True)
    _, entry, payload, called_method, body_method, url, is_delivery_context = candidates[0]
    return entry, payload, called_method, body_method, url, is_delivery_context


def pick_history_template(entries: list[dict[str, Any]]) -> tuple[dict[str, Any], dict[str, Any], str, str]:
    candidates: list[tuple[int, dict[str, Any], dict[str, Any], str, str]] = []

    for entry in entries:
        req = entry.get("request", {})
        called_method = get_header(req.get("headers", []), "x-calledmethod")
        if called_method != "Istoriya.History_Of_Instance":
            continue

        post_text = req.get("postData", {}).get("text")
        if not isinstance(post_text, str) or not post_text:
            continue

        try:
            payload = json.loads(post_text)
        except json.JSONDecodeError:
            continue

        params = payload.get("params")
        if not isinstance(params, dict):
            continue

        filter_record = params.get("Фильтр")
        if not isinstance(filter_record, dict):
            continue

        names = record_field_names(filter_record)
        if "GUID" not in names or "ИдО" not in names:
            continue

        score = 0
        if isinstance(params.get("Навигация"), dict):
            score += 2
        actions = get_record_field(filter_record, "Действие")
        if isinstance(actions, list) and "Изменение статуса заказа" in actions:
            score += 1

        candidates.append((score, entry, payload, called_method, payload.get("method", "История.History_Of_Instance")))

    if not candidates:
        raise RuntimeError("Не найден подходящий шаблон История.History_Of_Instance в HAR")

    candidates.sort(key=lambda item: item[0], reverse=True)
    _, entry, payload, called_method, body_method = candidates[0]
    return entry, payload, called_method, body_method


def try_pick_history_template(entries: list[dict[str, Any]]) -> tuple[dict[str, Any], dict[str, Any], str, str] | None:
    try:
        return pick_history_template(entries)
    except RuntimeError:
        return None


def load_history_template_from_fallback_har(active_har: Path) -> tuple[dict[str, Any], dict[str, Any], str, str] | None:
    for item in DEFAULT_HAR_CANDIDATES:
        candidate = Path(item).expanduser()
        if candidate.resolve() == active_har.resolve():
            continue
        if not candidate.exists():
            continue
        try:
            fallback_har = read_json(candidate)
            entries = fallback_har.get("log", {}).get("entries", [])
            if not isinstance(entries, list) or not entries:
                continue
            picked = try_pick_history_template(entries)
            if picked is not None:
                print(f"History-шаблон взят из резервного HAR: {candidate}")
                return picked
        except Exception:  # noqa: BLE001
            continue
    return None


def extract_position_meta(
    payload: dict[str, Any],
    default_field_name: str,
    default_field_type: str,
) -> tuple[str, str]:
    params = payload.get("params", {})
    nav = params.get("Навигация")
    if not isinstance(nav, dict):
        return default_field_name, default_field_type

    nav_values = nav.get("d", [])
    if not isinstance(nav_values, list) or len(nav_values) < 4:
        return default_field_name, default_field_type

    position = nav_values[3]
    if not isinstance(position, dict):
        return default_field_name, default_field_type

    specs = position.get("s", [])
    if not isinstance(specs, list) or not specs:
        return default_field_name, default_field_type

    first = specs[0]
    if not isinstance(first, dict):
        return default_field_name, default_field_type

    name = str(first.get("n", default_field_name))
    typ = str(first.get("t", default_field_type))
    return name or default_field_name, typ or default_field_type


def build_templates_from_har(har_data: dict[str, Any], active_har_path: Path | None = None) -> TemplateBundle:
    entries = har_data.get("log", {}).get("entries", [])
    if not isinstance(entries, list) or not entries:
        raise RuntimeError("HAR не содержит записей log.entries")

    (
        sale_entry,
        sale_payload,
        sale_called_method,
        sale_body_method,
        service_url,
        sale_is_delivery_context,
    ) = pick_sale_template(entries)
    history_picked = try_pick_history_template(entries)
    if history_picked is None and active_har_path is not None:
        history_picked = load_history_template_from_fallback_har(active_har_path)

    sale_payload["method"] = sale_body_method
    if history_picked is None:
        history_payload = make_default_history_payload_template()
        history_called_method = "Istoriya.History_Of_Instance"
    else:
        _, history_payload, history_called_method, history_body_method = history_picked
        history_payload["method"] = history_body_method

    sale_position_field, sale_position_type = extract_position_meta(
        sale_payload,
        default_field_name="NextDateWTZText",
        default_field_type="Строка",
    )
    history_position_field, history_position_type = extract_position_meta(
        history_payload,
        default_field_name="_time",
        default_field_type="Дата и время",
    )

    base_headers = headers_to_base(sale_entry.get("request", {}).get("headers", []))

    return TemplateBundle(
        service_url=service_url,
        base_headers=base_headers,
        sale_payload_template=sale_payload,
        sale_called_method=sale_called_method,
        sale_position_field=sale_position_field,
        sale_position_type=sale_position_type,
        har_sale_delivery_context=sale_is_delivery_context,
        history_payload_template=history_payload,
        history_called_method=history_called_method,
        history_position_field=history_position_field,
        history_position_type=history_position_type,
    )


def ensure_logged_in(page: Page) -> None:
    page.goto(DELIVERY_URL, wait_until="domcontentloaded")
    page.wait_for_timeout(3_000)

    if "rest.saby.ru/page/delivery" in page.url:
        return

    print("Открылся экран входа. Выполните вход вручную в окне браузера.")
    input("После успешного входа нажмите Enter...")

    page.goto(DELIVERY_URL, wait_until="domcontentloaded")
    page.wait_for_timeout(3_000)

    if "rest.saby.ru/page/delivery" not in page.url:
        raise RuntimeError(
            "Не удалось открыть страницу доставки после входа. Проверьте авторизацию и попробуйте снова."
        )


def wait_until_service_ready(
    clients: list[SabyRpcClient],
    template: TemplateBundle,
    target_date: date,
    page: Page,
) -> SabyRpcClient:
    position = build_position_record(
        template.sale_position_field,
        template.sale_position_type,
        f"{(target_date + timedelta(days=1)):%Y-%m-%d} 00:00:00.000000",
    )

    while True:
        payload = build_sale_payload(
            template=template,
            target_date=target_date,
            limit=1,
            direction="bothways",
            position=position,
        )

        errors: list[str] = []
        for client in clients:
            try:
                client.call(payload, template.sale_called_method)
                return client
            except Exception as err:  # noqa: BLE001
                errors.append(str(err))

        format_errors = [
            msg
            for msg in errors
            if (
                "Unable to parse value of field" in msg
                or "Тип не преобразуется" in msg
                or "Тип не преобразовывается" in msg
            )
        ]
        if format_errors:
            raise RuntimeError(
                "Ошибка формата фильтра SaleOrder.List. "
                "Проверьте runtime-контекст (несовместимые поля). "
                f"Пример ошибки: {format_errors[0]}"
            )

        method_errors = [
            msg
            for msg in errors
            if (
                "SaleOrder." in msg
                and ("не найден" in msg or "not found" in msg or "недоступен" in msg)
            )
        ]
        if method_errors:
            raise RuntimeError(
                "Сервер отклонил метод SaleOrder.List (ошибка метода/сигнатуры). "
                f"Пример ошибки: {method_errors[0]}"
            )

        cursor_errors = [
            msg
            for msg in errors
            if "Для текущего реестра разрешена курсорная навигация только по полям" in msg
        ]
        if cursor_errors:
            raise RuntimeError(
                "Сервер отклонил метод реестра заказов (ошибка метода/сигнатуры). "
                f"Пример ошибки: {cursor_errors[0]}"
            )

        print()
        print("Похоже, сессия не авторизована или пароль еще не введен.")
        print("Последние ошибки API:")
        for idx, msg in enumerate(errors, start=1):
            print(f"  {idx}. {msg}")
        if sys.stdin is not None and sys.stdin.isatty():
            print("Введите пароль/подтвердите вход в окне браузера и нажмите Enter.")
            try:
                input()
            except EOFError:
                time.sleep(3)
        else:
            print("Ожидаю авторизацию в окне браузера (авторежим, повтор через 3 сек)...")
            time.sleep(3)
        try:
            page.bring_to_front()
        except Exception:  # noqa: BLE001
            pass


def capture_runtime_service_meta(
    page: Page,
    timeout_seconds: int = 20,
) -> tuple[str | None, dict[str, str], dict[str, Any] | None, str | None, bool]:
    captures: list[dict[str, Any]] = []
    browser_context = page.context

    def _is_target_closed_error(err: Exception) -> bool:
        msg = str(err)
        return "Target page, context or browser has been closed" in msg

    def handle_request(request: Any) -> None:
        if request.method != "POST":
            return
        if "/service/" not in request.url:
            return
        headers = {str(k).lower(): str(v) for k, v in request.headers.items()}
        if "x-calledmethod" not in headers:
            return
        captures.append(
            {
                "url": request.url,
                "headers": headers,
                "called_method": headers.get("x-calledmethod", ""),
                "referer": headers.get("referer", ""),
                "post_data": request.post_data,
            }
        )

    def _is_list_like(called_method: str, payload: dict[str, Any] | None) -> bool:
        called = (called_method or "").lower()
        if not called.startswith("saleorder."):
            return False
        if called.startswith("saleorder.list") or ".list" in called:
            return True
        if not isinstance(payload, dict):
            return False
        params = payload.get("params")
        if not isinstance(params, dict):
            return False
        filter_record = params.get("Фильтр")
        if not isinstance(filter_record, dict):
            return False
        names = set(record_field_names(filter_record))
        # Для списка заказов ожидаем хотя бы часть датного фильтра.
        if {"DateTimeStartWTZ", "DateTimeEndWTZ"} & names:
            return True
        return False

    def _capture_is_list_like(c: dict[str, Any]) -> bool:
        called = str(c.get("called_method", ""))
        post_data = c.get("post_data")
        if not isinstance(post_data, str) or not post_data:
            return False
        try:
            payload = json.loads(post_data)
        except json.JSONDecodeError:
            return False
        return _is_list_like(called, payload if isinstance(payload, dict) else None)

    browser_context.on("request", handle_request)
    try:
        for _ in range(3):
            try:
                if page.is_closed():
                    page = browser_context.new_page()
                page.goto(DELIVERY_URL, wait_until="domcontentloaded")
                page.wait_for_timeout(700)
                # Принудительно активируем "Выполнен", чтобы runtime SaleOrder.List
                # был из нужного delivery-контекста.
                try:
                    page.get_by_text("Выполнен").first.click(timeout=2500)
                    page.wait_for_timeout(800)
                except Exception:  # noqa: BLE001
                    pass
                page.reload(wait_until="domcontentloaded")
                break
            except Exception as err:  # noqa: BLE001
                if not _is_target_closed_error(err):
                    raise
                try:
                    page = browser_context.new_page()
                except Exception:  # noqa: BLE001
                    return None, {}, None, None, False
        else:
            return None, {}, None, None, False
        deadline = time.time() + timeout_seconds
        while time.time() < deadline:
            if any(
                _capture_is_list_like(c)
                and "/page/delivery" in c.get("referer", "")
                for c in captures
            ):
                break
            try:
                if page.is_closed():
                    page = browser_context.new_page()
                    page.goto(DELIVERY_URL, wait_until="domcontentloaded")
                page.wait_for_timeout(250)
            except Exception as err:  # noqa: BLE001
                if not _is_target_closed_error(err):
                    raise
                try:
                    page = browser_context.new_page()
                    page.goto(DELIVERY_URL, wait_until="domcontentloaded")
                except Exception:  # noqa: BLE001
                    break
    finally:
        browser_context.remove_listener("request", handle_request)

    if not captures:
        return None, {}, None, None, False

    # Приоритетно берем запросы delivery-модуля, затем любые SaleOrder.*, затем последний.
    captures_sorted = sorted(
        captures,
        key=lambda c: (
            0
            if (c.get("called_method", "").startswith("SaleOrder.") and "/page/delivery" in c.get("referer", ""))
            else 1
            if c.get("called_method", "").startswith("SaleOrder.")
            else 2,
            -len(c.get("url", "")),
        ),
    )
    selected = next((c for c in captures_sorted if _capture_is_list_like(c)), captures_sorted[0])

    runtime_headers: dict[str, str] = {}
    headers = selected.get("headers", {})
    for key in ALLOWED_BASE_HEADERS:
        if key in headers and headers[key]:
            runtime_headers[key] = headers[key]

    runtime_sale_payload: dict[str, Any] | None = None
    runtime_sale_called_method: str | None = None
    runtime_sale_is_done = False

    for item in captures_sorted:
        called = str(item.get("called_method", ""))
        if not called.startswith("SaleOrder."):
            continue
        post_data = item.get("post_data")
        if not isinstance(post_data, str) or not post_data:
            continue
        try:
            parsed_payload = json.loads(post_data)
        except json.JSONDecodeError:
            continue
        if not (isinstance(parsed_payload, dict) and isinstance(parsed_payload.get("params"), dict)):
            continue
        if not _is_list_like(called, parsed_payload):
            continue
        if runtime_sale_payload is None:
            runtime_sale_payload = parsed_payload
            runtime_sale_called_method = called
        filter_record = parsed_payload.get("params", {}).get("Фильтр")
        filter_map = record_to_map(filter_record if isinstance(filter_record, dict) else None)
        if filter_map.get("ProductStateId") != "Done":
            continue
        runtime_sale_payload = parsed_payload
        runtime_sale_called_method = called
        runtime_sale_is_done = True
        break

    return (
        str(selected.get("url")),
        runtime_headers,
        runtime_sale_payload,
        runtime_sale_called_method,
        runtime_sale_is_done,
    )


def apply_runtime_sale_meta(
    templates: TemplateBundle,
    runtime_sale_payload: dict[str, Any] | None,
    runtime_sale_called_method: str | None,
    runtime_sale_is_done: bool,
    har_delivery_context: bool,
) -> None:
    if runtime_sale_payload and runtime_sale_called_method:
        runtime_method = runtime_sale_payload.get("method")
        if isinstance(runtime_method, str) and runtime_method:
            templates.sale_payload_template["method"] = runtime_method
        templates.sale_called_method = runtime_sale_called_method
        print(
            "Используются runtime-параметры SaleOrder.List: "
            f"method={templates.sale_payload_template.get('method')} "
            f"header={runtime_sale_called_method}"
        )
        if runtime_sale_is_done:
            merged_fields = merge_runtime_filter_context(
                templates.sale_payload_template,
                runtime_sale_payload,
            )
            print("Runtime-контекст: найден запрос ProductStateId=Done.")
            print(f"Runtime-контекст: перенесено полей фильтра: {merged_fields}")
        elif har_delivery_context:
            print(
                "Runtime-контекст: запрос ProductStateId=Done не найден, "
                "оставляю delivery-контекст из HAR (без переноса ближайших runtime-фильтров)."
            )
        else:
            merged_fields = merge_runtime_filter_context(
                templates.sale_payload_template,
                runtime_sale_payload,
            )
            print(
                "Runtime-контекст: запрос ProductStateId=Done не найден, "
                "взял контекст фильтров из ближайшего SaleOrder.List."
            )
            print(f"Runtime-контекст: перенесено полей фильтра: {merged_fields}")
    else:
        print(
            "Runtime-запрос SaleOrder.List не найден. Использую фильтр из HAR."
        )


def merge_service_url(template_url: str, runtime_url: str | None) -> str:
    if runtime_url:
        return runtime_url
    return template_url


def merge_headers(template_headers: dict[str, str], runtime_headers: dict[str, str]) -> dict[str, str]:
    out = dict(template_headers)
    out.update(runtime_headers)
    for key, value in DEFAULT_BASE_HEADERS.items():
        out.setdefault(key, value)
    return out


def build_position_record(field_name: str, field_type: str, value: Any) -> dict[str, Any]:
    return {
        "d": [value],
        "s": [{"t": field_type, "n": field_name}],
        "_type": "record",
        "f": 1,
    }


def build_navigation(
    nav_template: dict[str, Any] | None,
    direction: str,
    limit: int,
    position: dict[str, Any] | None,
) -> dict[str, Any]:
    nav = copy.deepcopy(nav_template) if isinstance(nav_template, dict) else {
        "d": [direction, True, limit, position],
        "s": [
            {"t": "Строка", "n": "Direction"},
            {"t": "Логическое", "n": "HasMore"},
            {"t": "Число целое", "n": "Limit"},
            {"t": "Запись", "n": "Position"},
        ],
        "_type": "record",
        "f": 0,
    }

    values = nav.setdefault("d", [])
    while len(values) < 4:
        values.append(None)
    values[0] = direction
    values[1] = True
    values[2] = limit
    values[3] = position

    # Некоторые шаблоны из HAR содержат Position как "Строка", но фактически API
    # ожидает "Запись" при постраничной навигации. Нормализуем схему.
    specs = nav.setdefault("s", [])
    if not isinstance(specs, list):
        specs = []
        nav["s"] = specs
    while len(specs) < 4:
        specs.append({})
    specs[0] = {"t": "Строка", "n": "Direction"}
    specs[1] = {"t": "Логическое", "n": "HasMore"}
    specs[2] = {"t": "Число целое", "n": "Limit"}
    specs[3] = {"t": "Запись", "n": "Position"}
    return nav


def ensure_sort_exists(params: dict[str, Any]) -> None:
    if params.get("Сортировка") is not None:
        return
    params["Сортировка"] = {
        "d": [[False, "NextDateWTZText", True]],
        "s": [
            {"t": "Логическое", "n": "l"},
            {"t": "Строка", "n": "n"},
            {"t": "Логическое", "n": "o"},
        ],
        "_type": "recordset",
        "f": 0,
    }


def recordset_to_dicts(recordset: dict[str, Any]) -> list[dict[str, Any]]:
    if recordset.get("_type") != "recordset":
        return []

    field_names = []
    for col in recordset.get("s", []):
        if isinstance(col, dict):
            field_names.append(str(col.get("n", "")))
        else:
            field_names.append(str(col))

    rows: list[dict[str, Any]] = []
    for row in recordset.get("d", []):
        if not isinstance(row, list):
            continue
        row_map: dict[str, Any] = {}
        for index, name in enumerate(field_names):
            row_map[name] = row[index] if index < len(row) else None
        rows.append(row_map)
    return rows


def find_navigation_record(result: dict[str, Any]) -> dict[str, Any] | None:
    candidates: list[dict[str, Any]] = []
    for key in ("Навигация", "navigation", "Navigation", "n"):
        value = result.get(key)
        if isinstance(value, dict):
            candidates.append(value)
    for value in result.values():
        if isinstance(value, dict):
            candidates.append(value)

    for nav in candidates:
        names = set(record_field_names(nav))
        if "Position" in names and ("HasMore" in names or "Direction" in names):
            return nav
    return None


def nav_has_more(nav: dict[str, Any] | None) -> bool | None:
    if not isinstance(nav, dict):
        return None
    for field_name in ("HasMore", "ЕстьЕще", "ЕстьЕщё"):
        value = get_record_field(nav, field_name)
        if isinstance(value, bool):
            return value
        if isinstance(value, int):
            return value != 0
    return None


def nav_position(nav: dict[str, Any] | None) -> dict[str, Any] | None:
    if not isinstance(nav, dict):
        return None
    value = get_record_field(nav, "Position")
    if isinstance(value, dict):
        return copy.deepcopy(value)
    return None


def position_signature(position: dict[str, Any] | None) -> str | None:
    if not isinstance(position, dict):
        return None
    try:
        return json.dumps(position, ensure_ascii=False, sort_keys=True)
    except Exception:  # noqa: BLE001
        return str(position)


def build_sale_payload(
    template: TemplateBundle,
    target_date: date,
    limit: int,
    direction: str,
    position: dict[str, Any] | None,
    reglament_override: int | None = None,
    clear_reglament: bool = False,
    context_relax_level: int = 0,
    scope_relax_mode: str = "keep",
) -> dict[str, Any]:
    payload = copy.deepcopy(template.sale_payload_template)
    payload["method"] = template.sale_payload_template.get("method", "SaleOrder.List")

    params = payload.get("params", {})
    filter_record = params.get("Фильтр")
    if not isinstance(filter_record, dict):
        raise RuntimeError("Некорректный шаблон: params.Фильтр отсутствует")

    start_text = f"{target_date:%Y-%m-%d} 00:00:00"
    end_text = f"{target_date:%Y-%m-%d} 23:59:59"

    set_record_field(filter_record, "DateTimeStartWTZ", start_text)
    set_record_field(filter_record, "DateTimeEndWTZ", end_text)
    set_record_field(filter_record, "ProductStateId", "Done")
    if clear_reglament:
        if not remove_record_field(filter_record, "Reglament"):
            set_record_field(filter_record, "Reglament", None)
    elif reglament_override is not None:
        set_record_field(filter_record, "Reglament", reglament_override)

    # Точечное расширение организационного охвата:
    # keep - как есть, no_ourorg - убираем только OurOrgFilter,
    # no_company - убираем только Company, no_org_scope - убираем оба.
    if scope_relax_mode in ("no_ourorg", "no_org_scope"):
        if not remove_record_field(filter_record, "OurOrgFilter"):
            set_record_field(filter_record, "OurOrgFilter", None)
    if scope_relax_mode in ("no_company", "no_org_scope"):
        if not remove_record_field(filter_record, "Company"):
            set_record_field(filter_record, "Company", None)

    if context_relax_level >= 1:
        # Снимаем возможный контекст CRM-фильтра (часто ограничивает до "своей" части заказов).
        if not remove_record_field(filter_record, "CRMFilter"):
            set_record_field(filter_record, "CRMFilter", None)
    if context_relax_level >= 2:
        # Дополнительно ослабляем контекст номенклатуры/полей для получения полного среза.
        for field_name in ("SaleNomenclatureFilter", "NotFields", "ProductStateFields"):
            if not remove_record_field(filter_record, field_name):
                set_record_field(filter_record, field_name, None)

    # Критично сохранять контекст delivery-фильтров из runtime/HAR:
    # принудительное обнуление Company/OurOrgFilter/Sales и т.п. уводит в заказы "Зала".
    # Поэтому базово не очищаем эти поля.

    # Для вкладки "Выполнен" в delivery обычно нужен подстатус 999.
    if record_field_index(filter_record, "ReglamentStates") is not None:
        set_record_field(filter_record, "ReglamentStates", [999])

    ensure_sort_exists(params)
    params["Навигация"] = build_navigation(params.get("Навигация"), direction, limit, position)
    return payload


def summarize_non_empty_filter_fields(filter_record: dict[str, Any]) -> list[str]:
    names = record_field_names(filter_record)
    values = filter_record.get("d", [])
    out: list[str] = []
    for idx, name in enumerate(names):
        value = values[idx] if idx < len(values) else None
        if value in (None, False, "", [], {}):
            continue
        if isinstance(value, (dict, list)):
            short = f"{type(value).__name__}(len={len(value)})"
        else:
            short = str(value)
        if len(short) > 120:
            short = short[:117] + "..."
        out.append(f"{name}={short}")
    return out


def build_history_payload(
    template: TemplateBundle,
    sale_id: int,
    sale_key: str,
    limit: int,
    position: dict[str, Any] | None,
) -> dict[str, Any]:
    payload = copy.deepcopy(template.history_payload_template)
    payload["method"] = template.history_payload_template.get("method", "История.History_Of_Instance")

    params = payload.get("params", {})
    filter_record = params.get("Фильтр")
    if not isinstance(filter_record, dict):
        raise RuntimeError("Некорректный шаблон: params.Фильтр отсутствует")

    if not set_record_field(filter_record, "GUID", sale_key):
        raise RuntimeError("Некорректный шаблон истории: поле GUID не найдено")
    if not set_record_field(filter_record, "ИдО", int(sale_id)):
        raise RuntimeError("Некорректный шаблон истории: поле ИдО не найдено")

    params["Навигация"] = build_navigation(params.get("Навигация"), "forward", limit, position)
    return payload


def fetch_done_orders(
    client: SabyRpcClient,
    template: TemplateBundle,
    target_date: date,
    page_limit: int,
    max_pages: int | None,
    reglament_override: int | None = None,
    clear_reglament: bool = False,
    verbose: bool = True,
    context_relax_level: int = 0,
    scope_relax_mode: str = "keep",
) -> list[dict[str, Any]]:
    orders: list[dict[str, Any]] = []
    seen_sale_ids: set[int] = set()

    first_position_value = f"{(target_date + timedelta(days=1)):%Y-%m-%d} 00:00:00.000000"
    position = build_position_record(template.sale_position_field, template.sale_position_type, first_position_value)

    direction = "bothways"
    page_number = 0

    while True:
        page_number += 1
        payload = build_sale_payload(
            template,
            target_date,
            page_limit,
            direction,
            position,
            reglament_override=reglament_override,
            clear_reglament=clear_reglament,
            context_relax_level=context_relax_level,
            scope_relax_mode=scope_relax_mode,
        )
        if page_number == 1:
            filter_record = payload.get("params", {}).get("Фильтр", {})
            if isinstance(filter_record, dict):
                active = summarize_non_empty_filter_fields(filter_record)
                if verbose:
                    print("[orders] активные фильтры:")
                    for item in active:
                        print(f"[orders]   {item}")
        result = client.call(payload, template.sale_called_method)
        rows = recordset_to_dicts(result)

        if not rows:
            break

        added = 0
        for row in rows:
            sale_id = row.get("Sale")
            if not isinstance(sale_id, int):
                continue
            if sale_id in seen_sale_ids:
                continue
            seen_sale_ids.add(sale_id)
            orders.append(row)
            added += 1

        if verbose:
            print(
                f"[orders] страница={page_number} строк={len(rows)} добавлено={added} всего={len(orders)}"
            )

        if max_pages is not None and page_number >= max_pages:
            break

        nav = find_navigation_record(result)
        has_more = nav_has_more(nav)
        next_position = nav_position(nav)

        if next_position is None:
            next_position_value = rows[-1].get(template.sale_position_field)
            if not isinstance(next_position_value, str) or not next_position_value:
                if has_more is False:
                    break
                if len(rows) < page_limit:
                    break
                break
            next_position = build_position_record(
                template.sale_position_field,
                template.sale_position_type,
                next_position_value,
            )

        if has_more is False and len(rows) < page_limit:
            break

        current_sig = position_signature(position)
        next_sig = position_signature(next_position)
        if next_sig is not None and next_sig == current_sig:
            break

        position = next_position
        direction = "forward"

    return orders


def extract_ui_done_count(page: Page) -> int | None:
    patterns = [
        re.compile(r"Выполнен\s*(\d{1,6})"),
        re.compile(r"Отмечено\s*(\d{1,6})"),
    ]
    for _ in range(8):
        try:
            text = page.inner_text("body")
        except Exception:  # noqa: BLE001
            text = ""
        if isinstance(text, str) and text:
            normalized = text.replace("\u00a0", " ")
            for pattern in patterns:
                match = pattern.search(normalized)
                if match:
                    try:
                        return int(match.group(1))
                    except ValueError:
                        pass
        page.wait_for_timeout(400)
    return None


def refine_orders_to_ui_count(
    orders: list[dict[str, Any]],
    target_count: int,
) -> tuple[list[dict[str, Any]], str | None]:
    if not orders or target_count <= 0:
        return orders, None

    fields = ("Company", "RealCompany", "OriginCompany", "Source", "Type", "Warehouse")
    best_subset = orders
    best_meta: str | None = None
    best_diff = abs(len(orders) - target_count)

    # Сначала пробуем срезы по одному полю.
    for field_name in fields:
        buckets: dict[Any, list[dict[str, Any]]] = {}
        for order in orders:
            value = order.get(field_name)
            if value is None:
                continue
            buckets.setdefault(value, []).append(order)
        if not buckets:
            continue

        candidates = sorted(
            ((value, bucket) for value, bucket in buckets.items()),
            key=lambda item: (
                abs(len(item[1]) - target_count),
                0 if len(item[1]) >= target_count else 1,
            ),
        )
        value, bucket = candidates[0]
        diff = abs(len(bucket) - target_count)
        if diff < best_diff:
            best_diff = diff
            best_subset = bucket
            best_meta = f"{field_name}={value} -> {len(bucket)}"

    # Затем пробуем срезы по комбинациям из двух полей.
    for idx_a, field_a in enumerate(fields):
        for field_b in fields[idx_a + 1 :]:
            buckets2: dict[tuple[Any, Any], list[dict[str, Any]]] = {}
            for order in orders:
                va = order.get(field_a)
                vb = order.get(field_b)
                if va is None or vb is None:
                    continue
                buckets2.setdefault((va, vb), []).append(order)
            if not buckets2:
                continue

            candidates2 = sorted(
                ((value_pair, bucket) for value_pair, bucket in buckets2.items()),
                key=lambda item: (
                    abs(len(item[1]) - target_count),
                    0 if len(item[1]) >= target_count else 1,
                ),
            )
            value_pair, bucket = candidates2[0]
            diff = abs(len(bucket) - target_count)
            if diff < best_diff:
                best_diff = diff
                best_subset = bucket
                best_meta = (
                    f"{field_a}={value_pair[0]}, {field_b}={value_pair[1]} "
                    f"-> {len(bucket)}"
                )

    # Наконец, пробуем комбинацию нескольких значений одного поля (subset-sum по бакетам).
    for field_name in fields:
        buckets: dict[Any, list[dict[str, Any]]] = {}
        for order in orders:
            value = order.get(field_name)
            if value is None:
                continue
            buckets.setdefault(value, []).append(order)
        if len(buckets) < 2:
            continue

        items = [(value, len(bucket)) for value, bucket in buckets.items() if len(bucket) > 0]
        if not items:
            continue
        # Ограничиваем перебор по верхней сумме около target.
        max_single = max(cnt for _, cnt in items)
        upper = max(target_count + max_single, target_count + 200)

        dp: dict[int, tuple[Any, ...]] = {0: ()}
        for value, cnt in items:
            snapshot = list(dp.items())
            for current_sum, values_tuple in snapshot:
                new_sum = current_sum + cnt
                if new_sum > upper:
                    continue
                if new_sum not in dp:
                    dp[new_sum] = values_tuple + (value,)

        if len(dp) <= 1:
            continue

        best_sum = min(
            (s for s in dp.keys() if s > 0),
            key=lambda s: (abs(s - target_count), 0 if s >= target_count else 1),
        )
        chosen_values = set(dp[best_sum])
        subset: list[dict[str, Any]] = []
        for value in chosen_values:
            subset.extend(buckets.get(value, []))
        diff = abs(len(subset) - target_count)
        if diff < best_diff:
            best_diff = diff
            best_subset = subset
            best_meta = (
                f"{field_name} in {sorted(chosen_values)} "
                f"-> {len(subset)} (values={len(chosen_values)})"
            )

    return best_subset, best_meta


def parse_status_change(message: str | None) -> tuple[str, str] | None:
    if not message:
        return None
    match = STATUS_CHANGE_PATTERN.search(message)
    if not match:
        return None
    return match.group("from"), match.group("to")


def is_history_method_not_found_error(message: str) -> bool:
    msg = (message or "").lower()
    return (
        "history_of_instance" in msg
        and (
            "не найден" in msg
            or "not found" in msg
            or "недоступен" in msg
            or "rpc http error 404" in msg
        )
    )


def history_method_variants(
    template: TemplateBundle,
    payload: dict[str, Any],
) -> list[tuple[str, str]]:
    current_called = str(template.history_called_method or "Istoriya.History_Of_Instance")
    current_body = str(payload.get("method") or "История.History_Of_Instance")
    variants: list[tuple[str, str]] = [
        (current_called, current_body),
        ("Istoriya.History_Of_Instance", "Istoriya.History_Of_Instance"),
        ("Istoriya.History_Of_Instance", "История.History_Of_Instance"),
        ("История.History_Of_Instance", "История.History_Of_Instance"),
        ("History.History_Of_Instance", "History.History_Of_Instance"),
    ]
    unique: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for item in variants:
        if item in seen:
            continue
        seen.add(item)
        unique.append(item)
    return unique


def call_history_with_auto_method(
    client: SabyRpcClient,
    template: TemplateBundle,
    payload: dict[str, Any],
) -> dict[str, Any]:
    last_err: Exception | None = None
    variants = history_method_variants(template, payload)
    for called_method, body_method in variants:
        trial_payload = copy.deepcopy(payload)
        trial_payload["method"] = body_method
        try:
            result = client.call(trial_payload, called_method)
            if called_method != template.history_called_method or body_method != str(
                template.history_payload_template.get("method")
            ):
                print(
                    "Автоподбор history-метода: "
                    f"header={called_method}, body={body_method}"
                )
            template.history_called_method = called_method
            template.history_payload_template["method"] = body_method
            return result
        except Exception as err:  # noqa: BLE001
            last_err = err
            if not is_history_method_not_found_error(str(err)):
                raise
            continue

    if last_err is not None:
        raise last_err
    raise RuntimeError("Не удалось выполнить history-запрос: нет рабочих вариантов метода.")


def fetch_order_status_history(
    client: SabyRpcClient,
    template: TemplateBundle,
    order: dict[str, Any],
    history_page_limit: int,
    max_history_pages: int | None,
) -> list[dict[str, Any]]:
    sale_id = order.get("Sale")
    sale_key = order.get("Key")
    if not isinstance(sale_id, int) or not isinstance(sale_key, str):
        return []

    statuses: list[dict[str, Any]] = []
    seen_events: set[tuple[Any, Any, Any]] = set()

    position: dict[str, Any] | None = None
    page_number = 0

    while True:
        page_number += 1
        payload = build_history_payload(
            template,
            sale_id=sale_id,
            sale_key=sale_key,
            limit=history_page_limit,
            position=position,
        )
        try:
            result = call_history_with_auto_method(client, template, payload)
        except Exception as err:  # noqa: BLE001
            print(
                f"Предупреждение: не удалось загрузить history для заказа {sale_id} ({order.get('Number')}): {err}"
            )
            return statuses
        events = recordset_to_dicts(result)

        if not events:
            break

        for event in events:
            event_key = (
                event.get("_event_id"),
                event.get("_time"),
                event.get("_message"),
            )
            if event_key in seen_events:
                continue
            seen_events.add(event_key)

            parsed = parse_status_change(event.get("_message"))
            if not parsed:
                continue

            from_status, to_status = parsed
            statuses.append(
                {
                    "Sale": sale_id,
                    "Key": sale_key,
                    "Number": order.get("Number"),
                    "DateWTZ": order.get("DateWTZ"),
                    "CustomerName": order.get("CustomerName"),
                    "StatusTime": event.get("_time") or event.get("server_time"),
                    "StatusFrom": from_status,
                    "StatusTo": to_status,
                    "Action": event.get("_action"),
                    "Message": event.get("_message"),
                }
            )

        if max_history_pages is not None and page_number >= max_history_pages:
            break

        nav = find_navigation_record(result)
        has_more = nav_has_more(nav)
        next_position = nav_position(nav)

        if next_position is None:
            next_position_value = events[-1].get(template.history_position_field) or events[-1].get("_time")
            if not isinstance(next_position_value, str) or not next_position_value:
                if has_more is False:
                    break
                if len(events) < history_page_limit:
                    break
                break
            next_position = build_position_record(
                template.history_position_field,
                template.history_position_type,
                next_position_value,
            )

        if has_more is False and len(events) < history_page_limit:
            break

        current_sig = position_signature(position)
        next_sig = position_signature(next_position)
        if next_sig is not None and next_sig == current_sig:
            break

        position = next_position

    return statuses


def autosize_worksheet_columns(ws: Any, max_width: int = 80) -> None:
    for column_index, column in enumerate(ws.columns, start=1):
        best = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > best:
                best = len(value)
        ws.column_dimensions[get_column_letter(column_index)].width = min(best + 2, max_width)


def parse_datetime_safe(value: Any) -> datetime | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    normalized = text.replace("T", " ").replace("Z", "")
    for fmt in (
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%d.%m.%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(normalized, fmt)
        except ValueError:
            continue
    try:
        dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if dt.tzinfo is not None:
            return dt.replace(tzinfo=None)
        return dt
    except ValueError:
        return None


def status_norm(value: Any) -> str:
    return str(value or "").strip().lower().replace("ё", "е")


def status_contains(value: Any, needles: tuple[str, ...]) -> bool:
    s = status_norm(value)
    return any(n in s for n in needles)


def minutes_between(start: datetime | None, end: datetime | None) -> float | None:
    if not start or not end:
        return None
    if end < start:
        return None
    return (end - start).total_seconds() / 60.0


def percentile(values: list[float], q: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    idx = max(0, min(len(ordered) - 1, math.ceil(q * len(ordered)) - 1))
    return ordered[idx]


def pick_first_non_empty(row: dict[str, Any], keys: tuple[str, ...]) -> Any:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return None


def build_delivery_analytics(
    target_date: date,
    orders: list[dict[str, Any]],
    statuses: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    status_by_sale: dict[int, list[dict[str, Any]]] = {}
    for item in statuses:
        sale = item.get("Sale")
        if isinstance(sale, int):
            status_by_sale.setdefault(sale, []).append(item)

    for events in status_by_sale.values():
        events.sort(key=lambda x: x.get("StatusTime") or "")

    order_metrics: list[dict[str, Any]] = []
    stage_rows: list[dict[str, Any]] = []
    restaurant_rows: list[dict[str, Any]] = []

    for order in orders:
        sale = order.get("Sale")
        if not isinstance(sale, int):
            continue

        events = status_by_sale.get(sale, [])
        first_status_time = parse_datetime_safe(events[0].get("StatusTime")) if events else None
        order_start = parse_datetime_safe(order.get("DateWTZ")) or first_status_time

        t_cooking: datetime | None = None
        t_assembly: datetime | None = None
        t_delivery: datetime | None = None
        t_done: datetime | None = None

        for ev in events:
            t = parse_datetime_safe(ev.get("StatusTime"))
            to_status = ev.get("StatusTo")
            if t_cooking is None and status_contains(to_status, ("готов",)):
                t_cooking = t
            if t_assembly is None and status_contains(to_status, ("сборк",)):
                t_assembly = t
            if t_delivery is None and status_contains(to_status, ("доставк", "в пути", "у курьер")):
                t_delivery = t
            if t_done is None and status_contains(to_status, ("выполн", "заверш", "закрыт")):
                t_done = t

        t_done = t_done or parse_datetime_safe(order.get("ClosedWTZ"))

        processing_min = minutes_between(order_start, t_cooking)
        cooking_min = minutes_between(t_cooking, t_assembly)
        assembly_min = minutes_between(t_assembly, t_delivery)
        delivery_min = minutes_between(t_delivery, t_done)
        total_min = minutes_between(order_start, t_done)

        stages = {
            "Обработка": processing_min,
            "Готовка": cooking_min,
            "Сборка": assembly_min,
            "Доставка": delivery_min,
        }
        known = {k: v for k, v in stages.items() if isinstance(v, (int, float))}
        bottleneck_stage = None
        bottleneck_min = None
        if known:
            bottleneck_stage, bottleneck_min = max(known.items(), key=lambda x: x[1])

        restaurant = pick_first_non_empty(
            order,
            (
                "WarehouseName",
                "CompanyName",
                "RealCompanyName",
                "OriginCompanyName",
                "SalesPointName",
            ),
        )
        courier = pick_first_non_empty(order, ("CourierName", "Courier", "CourierFIO", "CourierEmployee"))
        operator = pick_first_non_empty(
            order,
            (
                "OperatorName",
                "Operator",
                "ManagerName",
                "EmployeeName",
                "ResponsibleName",
                "CreatedByName",
            ),
        )

        order_metrics.append(
            {
                "ScanDate": f"{target_date:%Y-%m-%d}",
                "Sale": sale,
                "Number": order.get("Number"),
                "Restaurant": restaurant,
                "Courier": courier,
                "Operator": operator,
                "OrderStart": order_start,
                "DoneTime": t_done,
                "TotalMin": total_min,
                "ProcessingMin": processing_min,
                "CookingMin": cooking_min,
                "AssemblyMin": assembly_min,
                "DeliveryMin": delivery_min,
                "BottleneckStage": bottleneck_stage,
                "BottleneckMin": bottleneck_min,
            }
        )

    def collect(values_key: str) -> list[float]:
        return [float(x[values_key]) for x in order_metrics if isinstance(x.get(values_key), (int, float))]

    for stage_name, key in (
        ("Обработка", "ProcessingMin"),
        ("Готовка", "CookingMin"),
        ("Сборка", "AssemblyMin"),
        ("Доставка", "DeliveryMin"),
        ("Итого", "TotalMin"),
    ):
        vals = collect(key)
        if not vals:
            continue
        stage_rows.append(
            {
                "Metric": stage_name,
                "Count": len(vals),
                "AvgMin": statistics.mean(vals),
                "MedianMin": statistics.median(vals),
                "P90Min": percentile(vals, 0.9),
                "MaxMin": max(vals),
            }
        )

    by_restaurant: dict[str, list[dict[str, Any]]] = {}
    for row in order_metrics:
        name = str(row.get("Restaurant") or "Не указан")
        by_restaurant.setdefault(name, []).append(row)

    for name, items in sorted(by_restaurant.items(), key=lambda kv: len(kv[1]), reverse=True):
        total_vals = [float(x["TotalMin"]) for x in items if isinstance(x.get("TotalMin"), (int, float))]
        deliv_vals = [float(x["DeliveryMin"]) for x in items if isinstance(x.get("DeliveryMin"), (int, float))]
        restaurant_rows.append(
            {
                "Restaurant": name,
                "Orders": len(items),
                "AvgTotalMin": statistics.mean(total_vals) if total_vals else None,
                "AvgDeliveryMin": statistics.mean(deliv_vals) if deliv_vals else None,
                "P90TotalMin": percentile(total_vals, 0.9) if total_vals else None,
            }
        )

    return order_metrics, stage_rows, restaurant_rows


def build_problem_orders_rows(order_metrics: list[dict[str, Any]], top_n: int = 20) -> list[dict[str, Any]]:
    def top_by(metric_key: str) -> list[dict[str, Any]]:
        rows = [row for row in order_metrics if isinstance(row.get(metric_key), (int, float))]
        rows.sort(key=lambda x: float(x.get(metric_key) or 0.0), reverse=True)
        return rows[:top_n]

    out: list[dict[str, Any]] = []

    top_delivery = top_by("DeliveryMin")
    for idx, row in enumerate(top_delivery, start=1):
        out.append(
            {
                "Metric": "DeliveryMin",
                "Rank": idx,
                **row,
            }
        )

    top_total = top_by("TotalMin")
    for idx, row in enumerate(top_total, start=1):
        out.append(
            {
                "Metric": "TotalMin",
                "Rank": idx,
                **row,
            }
        )

    return out


def export_excel(
    output_path: Path,
    target_date: date,
    orders: list[dict[str, Any]],
    statuses: list[dict[str, Any]],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    registry_headers = [
        "ScanDate",
        "Sale",
        "Key",
        "Number",
        "DateWTZ",
        "ClosedWTZ",
        "CustomerName",
        "Customer",
        "CourierName",
        "ProductState",
        "ReglamentState",
        "Reglament",
        "Source",
        "Type",
        "Address",
        "WarehouseName",
        "TotalPrice",
        "NextDateWTZText",
    ]
    ws_registry = wb.active
    ws_registry.title = "Реестр"
    ws_registry.append(registry_headers)

    for order in orders:
        ws_registry.append(
            [
                f"{target_date:%Y-%m-%d}",
                order.get("Sale"),
                order.get("Key"),
                order.get("Number"),
                order.get("DateWTZ"),
                order.get("ClosedWTZ"),
                order.get("CustomerName"),
                order.get("Customer"),
                order.get("CourierName"),
                order.get("ProductState"),
                order.get("ReglamentState"),
                order.get("Reglament"),
                order.get("Source"),
                order.get("Type"),
                order.get("Address"),
                order.get("WarehouseName"),
                order.get("TotalPrice"),
                order.get("NextDateWTZText"),
            ]
        )

    status_headers = [
        "ScanDate",
        "Sale",
        "Key",
        "Number",
        "OrderDateWTZ",
        "CustomerName",
        "StatusTime",
        "StatusFrom",
        "StatusTo",
        "Action",
        "Message",
    ]
    ws_status = wb.create_sheet("Статусы")
    ws_status.append(status_headers)

    for row in sorted(
        statuses,
        key=lambda item: (
            item.get("Sale") or 0,
            item.get("StatusTime") or "",
        ),
    ):
        ws_status.append(
            [
                f"{target_date:%Y-%m-%d}",
                row.get("Sale"),
                row.get("Key"),
                row.get("Number"),
                row.get("DateWTZ"),
                row.get("CustomerName"),
                row.get("StatusTime"),
                row.get("StatusFrom"),
                row.get("StatusTo"),
                row.get("Action"),
                row.get("Message"),
            ]
        )

    autosize_worksheet_columns(ws_registry)
    autosize_worksheet_columns(ws_status)

    analytics_rows, stage_rows, restaurant_rows = build_delivery_analytics(
        target_date=target_date,
        orders=orders,
        statuses=statuses,
    )

    ws_analytics = wb.create_sheet("Аналитика")
    analytics_headers = [
        "ScanDate",
        "Sale",
        "Number",
        "Restaurant",
        "Courier",
        "Operator",
        "OrderStart",
        "DoneTime",
        "TotalMin",
        "ProcessingMin",
        "CookingMin",
        "AssemblyMin",
        "DeliveryMin",
        "BottleneckStage",
        "BottleneckMin",
    ]
    ws_analytics.append(analytics_headers)
    for row in analytics_rows:
        ws_analytics.append(
            [
                row.get("ScanDate"),
                row.get("Sale"),
                row.get("Number"),
                row.get("Restaurant"),
                row.get("Courier"),
                row.get("Operator"),
                row.get("OrderStart"),
                row.get("DoneTime"),
                row.get("TotalMin"),
                row.get("ProcessingMin"),
                row.get("CookingMin"),
                row.get("AssemblyMin"),
                row.get("DeliveryMin"),
                row.get("BottleneckStage"),
                row.get("BottleneckMin"),
            ]
        )

    ws_bottlenecks = wb.create_sheet("Узкие места")
    ws_bottlenecks.append(["Этап", "Кол-во", "Среднее, мин", "Медиана, мин", "P90, мин", "Макс, мин"])
    for row in stage_rows:
        ws_bottlenecks.append(
            [
                row.get("Metric"),
                row.get("Count"),
                row.get("AvgMin"),
                row.get("MedianMin"),
                row.get("P90Min"),
                row.get("MaxMin"),
            ]
        )

    ws_bottlenecks.append([])
    ws_bottlenecks.append(["Ресторан", "Заказы", "Среднее общее, мин", "Среднее доставка, мин", "P90 общее, мин"])
    for row in restaurant_rows:
        ws_bottlenecks.append(
            [
                row.get("Restaurant"),
                row.get("Orders"),
                row.get("AvgTotalMin"),
                row.get("AvgDeliveryMin"),
                row.get("P90TotalMin"),
            ]
        )

    problem_rows = build_problem_orders_rows(analytics_rows, top_n=20)
    ws_problems = wb.create_sheet("Проблемные заказы")
    ws_problems.append(
        [
            "Metric",
            "Rank",
            "Sale",
            "Number",
            "Restaurant",
            "Courier",
            "Operator",
            "TotalMin",
            "DeliveryMin",
            "ProcessingMin",
            "CookingMin",
            "AssemblyMin",
            "BottleneckStage",
            "BottleneckMin",
        ]
    )
    for row in problem_rows:
        ws_problems.append(
            [
                row.get("Metric"),
                row.get("Rank"),
                row.get("Sale"),
                row.get("Number"),
                row.get("Restaurant"),
                row.get("Courier"),
                row.get("Operator"),
                row.get("TotalMin"),
                row.get("DeliveryMin"),
                row.get("ProcessingMin"),
                row.get("CookingMin"),
                row.get("AssemblyMin"),
                row.get("BottleneckStage"),
                row.get("BottleneckMin"),
            ]
        )

    autosize_worksheet_columns(ws_analytics)
    autosize_worksheet_columns(ws_bottlenecks)
    autosize_worksheet_columns(ws_problems)

    wb.save(output_path)


def get_output_path(target_date: date, explicit_output: str | None) -> Path:
    if explicit_output:
        return Path(explicit_output).expanduser().resolve()
    return (Path.cwd() / "exports" / f"order_status_history_{target_date:%Y-%m-%d}.xlsx").resolve()


def is_profile_lock_error(err: Exception) -> bool:
    message = str(err)
    return (
        "ProcessSingleton" in message
        or "SingletonLock" in message
        or "profile is already in use" in message
    )


def is_playwright_browser_missing_error(err: Exception) -> bool:
    message = str(err)
    return (
        "Executable doesn't exist" in message
        or "Please run the following command to download new browsers" in message
        or "playwright install" in message
    )


def find_system_chromium_executable() -> Path | None:
    candidates: list[Path] = []
    if os.name == "nt":
        env_paths = [
            os.environ.get("PROGRAMFILES"),
            os.environ.get("PROGRAMFILES(X86)"),
            os.environ.get("LOCALAPPDATA"),
        ]
        for base in env_paths:
            if not base:
                continue
            root = Path(base)
            candidates.extend(
                [
                    root / "Google" / "Chrome" / "Application" / "chrome.exe",
                    root / "Microsoft" / "Edge" / "Application" / "msedge.exe",
                    root / "Chromium" / "Application" / "chrome.exe",
                ]
            )
    else:
        candidates.extend(
            [
                Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
                Path("/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge"),
                Path("/usr/bin/google-chrome"),
                Path("/usr/bin/chromium"),
                Path("/usr/bin/chromium-browser"),
            ]
        )
    for path in candidates:
        if path.exists():
            return path
    return None


def create_fallback_profile_dir(base_profile_dir: Path) -> Path:
    runs_dir = base_profile_dir.parent / ".saby_profile_runs"
    runs_dir.mkdir(parents=True, exist_ok=True)
    suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = runs_dir / f"profile_{suffix}"
    candidate.mkdir(parents=True, exist_ok=True)
    return candidate


def launch_context_with_profile_fallback(
    playwright_obj: Any,
    base_profile_dir: Path,
    headless: bool,
) -> tuple[BrowserContext, Path]:
    launch_kwargs = {
        "headless": bool(headless),
        "viewport": {"width": 1400, "height": 900},
        "locale": "ru-RU",
    }

    def _launch(user_data_dir: Path, executable_path: str | None = None) -> BrowserContext:
        kwargs = dict(launch_kwargs)
        if executable_path:
            kwargs["executable_path"] = executable_path
        return playwright_obj.chromium.launch_persistent_context(
            user_data_dir=str(user_data_dir),
            **kwargs,
        )

    executable_path: str | None = None
    try:
        context = _launch(base_profile_dir)
        return context, base_profile_dir
    except Exception as err:  # noqa: BLE001
        if is_playwright_browser_missing_error(err):
            system_browser = find_system_chromium_executable()
            if system_browser is not None:
                executable_path = str(system_browser)
                print(
                    "Встроенный браузер Playwright не найден. "
                    f"Использую системный Chromium-браузер: {system_browser}"
                )
                context = _launch(base_profile_dir, executable_path=executable_path)
                return context, base_profile_dir
        if not is_profile_lock_error(err):
            raise

        fallback_dir = create_fallback_profile_dir(base_profile_dir)
        print(
            "Основной профиль занят другим процессом Chromium. "
            f"Переключаюсь на временный профиль: {fallback_dir}"
        )
        context = _launch(fallback_dir, executable_path=executable_path)
        return context, fallback_dir


def main() -> int:
    args = parse_args()

    try:
        target_date = parse_date(args.date)
    except ValueError as err:
        print(f"Ошибка: {err}")
        return 2

    har_path = resolve_har_path(args.har)
    templates: TemplateBundle | None = None
    har_error: str | None = None
    if har_path:
        try:
            har_data = read_json(har_path)
            templates = build_templates_from_har(har_data, active_har_path=har_path)
        except Exception as err:  # noqa: BLE001
            har_error = str(err)
    else:
        har_error = (
            "HAR файл не найден. Проверены варианты: "
            + ", ".join(DEFAULT_HAR_CANDIDATES)
        )

    output_path = get_output_path(target_date, args.output)

    profile_dir = Path(args.profile_dir).expanduser().resolve()
    profile_dir.mkdir(parents=True, exist_ok=True)

    print(f"Дата сканирования: {target_date:%Y-%m-%d}")
    if templates is not None and har_path is not None:
        print(f"HAR шаблон: {har_path}")
    else:
        print("HAR шаблон: не используется (runtime fallback)")
        if har_error:
            print(f"Предупреждение HAR: {har_error}")
    print(f"Профиль браузера: {profile_dir}")

    with sync_playwright() as p:
        context, active_profile_dir = launch_context_with_profile_fallback(
            playwright_obj=p,
            base_profile_dir=profile_dir,
            headless=bool(args.headless),
        )
        try:
            if active_profile_dir != profile_dir:
                print(f"Активный профиль браузера: {active_profile_dir}")
            page = context.pages[0] if context.pages else context.new_page()
            ensure_logged_in(page)

            (
                runtime_service_url,
                runtime_headers,
                runtime_sale_payload,
                runtime_sale_called_method,
                runtime_sale_is_done,
            ) = capture_runtime_service_meta(page)
            if templates is None:
                templates = make_runtime_fallback_templates(
                    runtime_sale_payload=runtime_sale_payload,
                    runtime_sale_called_method=runtime_sale_called_method,
                )
                print(
                    "HAR недоступен/неподходящий. Перехожу в полностью автоматический runtime-режим."
                )
                if runtime_sale_payload is None:
                    print(
                        "Runtime-шаблон SaleOrder.List не пойман. "
                        "Использую встроенный универсальный шаблон SaleOrder.List."
                    )

            service_url = merge_service_url(templates.service_url, runtime_service_url)
            base_headers = merge_headers(templates.base_headers, runtime_headers)
            har_delivery_context = bool(templates.har_sale_delivery_context)
            apply_runtime_sale_meta(
                templates=templates,
                runtime_sale_payload=runtime_sale_payload,
                runtime_sale_called_method=runtime_sale_called_method,
                runtime_sale_is_done=runtime_sale_is_done,
                har_delivery_context=har_delivery_context,
            )

            parsed_url = urlparse(service_url)
            x_version = parse_qs(parsed_url.query).get("x_version", [None])[0]
            if x_version:
                print(f"Используем x_version: {x_version}")

            primary_client = SabyRpcClient(context, service_url, base_headers)
            fallback_client = SabyRpcClient(context, templates.service_url, templates.base_headers)
            try:
                client = wait_until_service_ready(
                    clients=[primary_client, fallback_client],
                    template=templates,
                    target_date=target_date,
                    page=page,
                )
            except RuntimeError as err:
                msg = str(err)
                if not (
                    "SaleOrder." in msg
                    or "метода/сигнатуры" in msg
                    or "курсорная навигация" in msg
                ):
                    raise
                print(
                    "Автовосстановление: переполучаю runtime-параметры реестра после ошибки метода..."
                )
                (
                    runtime_service_url,
                    runtime_headers,
                    runtime_sale_payload,
                    runtime_sale_called_method,
                    runtime_sale_is_done,
                ) = capture_runtime_service_meta(page, timeout_seconds=35)
                apply_runtime_sale_meta(
                    templates=templates,
                    runtime_sale_payload=runtime_sale_payload,
                    runtime_sale_called_method=runtime_sale_called_method,
                    runtime_sale_is_done=runtime_sale_is_done,
                    har_delivery_context=har_delivery_context,
                )
                service_url = merge_service_url(templates.service_url, runtime_service_url)
                base_headers = merge_headers(templates.base_headers, runtime_headers)
                primary_client = SabyRpcClient(context, service_url, base_headers)
                fallback_client = SabyRpcClient(context, templates.service_url, templates.base_headers)
                client = wait_until_service_ready(
                    clients=[primary_client, fallback_client],
                    template=templates,
                    target_date=target_date,
                    page=page,
                )

            ui_done_count: int | None = None
            if args.align_to_ui_count:
                ui_done_count = extract_ui_done_count(page)
                if ui_done_count is not None:
                    print(f"Счетчик UI (Выполнен/Отмечено): {ui_done_count}")
                else:
                    print("Не удалось прочитать счетчик UI (Выполнен/Отмечено).")
            else:
                print("UI-счетчик отключен (--align-to-ui-count не задан).")

            template_filter = templates.sale_payload_template.get("params", {}).get("Фильтр")
            template_reglament = None
            if isinstance(template_filter, dict):
                template_reglament = record_to_map(template_filter).get("Reglament")

            orders: list[dict[str, Any]] = []
            heuristic_fallback_enabled = bool(args.allow_heuristic_fallback)
            if not runtime_sale_is_done and not heuristic_fallback_enabled:
                heuristic_fallback_enabled = True
                print(
                    "Runtime Done-запрос не пойман. Автоматически включаю эвристический fallback."
                )
            if not runtime_sale_is_done and har_delivery_context and not heuristic_fallback_enabled:
                print(
                    "Runtime Done-запрос не пойман, но HAR содержит delivery-контекст SaleOrder.List. "
                    "Продолжаю в строгом HAR-режиме без эвристики."
                )
            elif not runtime_sale_is_done and not heuristic_fallback_enabled:
                raise RuntimeError(
                    "Не удалось поймать runtime-запрос SaleOrder.List с ProductStateId=Done. "
                    "Чтобы получить корректный состав заказов доставки, перед запуском откройте страницу "
                    "доставки, выберите дату и вкладку 'Выполнен', затем повторите запуск. "
                    "Если хотите принудительно использовать эвристику, запустите с флагом "
                    "--allow-heuristic-fallback."
                )

            ui_alignment_allowed = runtime_sale_is_done or heuristic_fallback_enabled

            if runtime_sale_is_done or ui_done_count is None or not heuristic_fallback_enabled:
                print("Загружаю реестр выполненных заказов...")
                orders = fetch_done_orders(
                    client=client,
                    template=templates,
                    target_date=target_date,
                    page_limit=int(args.order_page_limit),
                    max_pages=args.max_order_pages,
                    reglament_override=template_reglament if isinstance(template_reglament, int) else None,
                    clear_reglament=template_reglament is None,
                    verbose=True,
                    context_relax_level=0,
                    scope_relax_mode="keep",
                )
                if ui_done_count is not None and not ui_alignment_allowed:
                    print(
                        "UI-счетчик не используется для автоподбора фильтров: "
                        "runtime Done-запрос не пойман (строгий HAR-режим)."
                    )

                if ui_done_count is not None and ui_alignment_allowed:
                    diff = abs(len(orders) - ui_done_count)
                    if diff > max(10, int(ui_done_count * 0.02)):
                        print(
                            "Пробую варианты поля Reglament..."
                        )
                        reg_variants: list[tuple[str, int | None, bool]] = []
                        if isinstance(template_reglament, int):
                            reg_variants.append((f"Reglament={template_reglament}", template_reglament, False))
                        # В строгом delivery-режиме не используем Reglament=1:
                        # по бизнес-правилу это признак заказов зала.
                        for reg_value in (2,):
                            if not any((v == reg_value and not clear) for _, v, clear in reg_variants):
                                reg_variants.append((f"Reglament={reg_value}", reg_value, False))
                        reg_variants.append(("Reglament=<unset>", None, True))

                        reg_results: list[tuple[int, str, int | None, bool, list[dict[str, Any]]]] = []
                        for label, reg_value, clear_reg in reg_variants:
                            try:
                                reg_orders = fetch_done_orders(
                                    client=client,
                                    template=templates,
                                    target_date=target_date,
                                    page_limit=int(args.order_page_limit),
                                    max_pages=args.max_order_pages,
                                    reglament_override=reg_value,
                                    clear_reglament=clear_reg,
                                    verbose=False,
                                    context_relax_level=0,
                                    scope_relax_mode="keep",
                                )
                                print(f"Вариант {label}: найдено {len(reg_orders)}")
                                reg_results.append((len(reg_orders), label, reg_value, clear_reg, reg_orders))
                            except Exception as err:  # noqa: BLE001
                                print(f"Вариант {label}: ошибка ({err})")

                        if reg_results:
                            reg_results.sort(
                                key=lambda item: (
                                    abs(item[0] - ui_done_count),
                                    0 if item[0] >= ui_done_count else 1,
                                    item[0],
                                )
                            )
                            best_count, best_label, best_reg, best_clear_reg, best_orders = reg_results[0]
                            if best_count != len(orders):
                                print(f"Выбран вариант {best_label}: {best_count} (UI={ui_done_count})")
                                orders = best_orders
                                template_reglament = None if best_clear_reg else best_reg
                                diff = abs(len(orders) - ui_done_count)

                    if diff > max(10, int(ui_done_count * 0.02)):
                        print(
                            "Пробую варианты context_relax_level в строгом delivery-HAR режиме..."
                        )
                        relax_results: list[tuple[int, int, list[dict[str, Any]]]] = []
                        for relax_level in (0, 1, 2):
                            try:
                                relaxed_orders = fetch_done_orders(
                                    client=client,
                                    template=templates,
                                    target_date=target_date,
                                    page_limit=int(args.order_page_limit),
                                    max_pages=args.max_order_pages,
                                    reglament_override=template_reglament if isinstance(template_reglament, int) else None,
                                    clear_reglament=template_reglament is None,
                                    verbose=False,
                                    context_relax_level=relax_level,
                                    scope_relax_mode="keep",
                                )
                                print(
                                    f"Вариант context_relax_level={relax_level}: найдено {len(relaxed_orders)}"
                                )
                                relax_results.append((len(relaxed_orders), relax_level, relaxed_orders))
                            except Exception as err:  # noqa: BLE001
                                print(f"Вариант context_relax_level={relax_level}: ошибка ({err})")

                        if relax_results:
                            relax_results.sort(
                                key=lambda item: (
                                    abs(item[0] - ui_done_count),
                                    0 if item[0] >= ui_done_count else 1,
                                    item[0],
                                    item[1],
                                )
                            )
                            best_relaxed_count, best_relax_level, best_relaxed_orders = relax_results[0]
                            if best_relaxed_count != len(orders):
                                print(
                                    f"Выбран вариант context_relax_level={best_relax_level}: "
                                    f"{best_relaxed_count} (UI={ui_done_count})"
                                )
                                orders = best_relaxed_orders
                                diff = abs(len(orders) - ui_done_count)

                    if diff > max(10, int(ui_done_count * 0.02)):
                        print(
                            "Предупреждение: состав отличается от UI-счетчика "
                            f"(UI={ui_done_count}, выгрузка={len(orders)}). "
                            "Пробую расширить только организационный охват..."
                        )
                        scope_modes = ("keep", "no_ourorg", "no_company", "no_org_scope")
                        scope_variants: list[tuple[int, str, list[dict[str, Any]]]] = [(len(orders), "keep", orders)]
                        for scope_mode in scope_modes[1:]:
                            try:
                                scoped_orders = fetch_done_orders(
                                    client=client,
                                    template=templates,
                                    target_date=target_date,
                                    page_limit=int(args.order_page_limit),
                                    max_pages=args.max_order_pages,
                                    reglament_override=template_reglament if isinstance(template_reglament, int) else None,
                                    clear_reglament=template_reglament is None,
                                    verbose=False,
                                    context_relax_level=0,
                                    scope_relax_mode=scope_mode,
                                )
                                print(
                                    f"Вариант scope_relax_mode={scope_mode}: найдено {len(scoped_orders)}"
                                )
                                scope_variants.append((len(scoped_orders), scope_mode, scoped_orders))
                            except Exception as err:  # noqa: BLE001
                                print(f"Вариант scope_relax_mode={scope_mode}: ошибка ({err})")

                        scope_variants.sort(
                            key=lambda item: (
                                abs(item[0] - ui_done_count),
                                0 if item[0] >= ui_done_count else 1,
                            )
                        )
                        best_count, best_scope_mode, best_scope_orders = scope_variants[0]
                        if best_scope_mode != "keep":
                            print(
                                f"Выбран вариант scope_relax_mode={best_scope_mode}: "
                                f"{best_count} (UI={ui_done_count})"
                            )
                            orders = best_scope_orders
            else:
                candidate_reglaments: list[int | None] = []
                for candidate in (template_reglament, 1, 2):
                    if candidate not in candidate_reglaments:
                        candidate_reglaments.append(candidate)

                variants: list[tuple[int, int | None, int, list[dict[str, Any]]]] = []
                for candidate in candidate_reglaments:
                    for relax_level in (0, 1, 2):
                        print(
                            f"Пробую вариант фильтра: Reglament={candidate}, "
                            f"context_relax_level={relax_level}"
                        )
                        try:
                            probe_orders = fetch_done_orders(
                                client=client,
                                template=templates,
                                target_date=target_date,
                                page_limit=int(args.order_page_limit),
                                max_pages=args.max_order_pages,
                                reglament_override=candidate if isinstance(candidate, int) else None,
                                verbose=False,
                                context_relax_level=relax_level,
                            )
                            count = len(probe_orders)
                            print(
                                f"Вариант Reglament={candidate}, context_relax_level={relax_level}: "
                                f"найдено {count}"
                            )
                            variants.append(
                                (
                                    count,
                                    candidate if isinstance(candidate, int) else None,
                                    relax_level,
                                    probe_orders,
                                )
                            )
                        except Exception as err:  # noqa: BLE001
                            print(
                                f"Вариант Reglament={candidate}, context_relax_level={relax_level}: "
                                f"ошибка ({err})"
                            )

                # Приоритет:
                # 1) минимальная разница с UI;
                # 2) если разница одинаковая — берем вариант не меньше UI (чтобы не терять заказы);
                # 3) затем минимальный relax_level.
                if not variants:
                    raise RuntimeError("Не удалось получить ни одного варианта фильтра")
                variants.sort(
                    key=lambda item: (
                        abs(item[0] - ui_done_count),
                        0 if item[0] >= ui_done_count else 1,
                        item[2],
                    )
                )
                best_count, best_reglament, best_relax, best_orders = variants[0]
                print(
                    f"Выбран вариант Reglament={best_reglament}: "
                    f"{best_count} заказов (ближе всего к UI={ui_done_count}, "
                    f"context_relax_level={best_relax})"
                )
                orders = best_orders

                # Если даже лучший вариант заметно расходится с UI-счетчиком,
                # пытаемся найти устойчивый срез по ключевым полям заказа.
                current_diff = abs(len(orders) - ui_done_count)
                if current_diff > max(25, int(ui_done_count * 0.05)):
                    over_variants = [v for v in variants if v[0] >= ui_done_count]
                    source_for_refine = orders
                    if over_variants:
                        over_variants.sort(key=lambda item: (item[0] - ui_done_count, item[2]))
                        source_for_refine = over_variants[0][3]
                    refined_orders, refine_meta = refine_orders_to_ui_count(source_for_refine, ui_done_count)
                    refined_diff = abs(len(refined_orders) - ui_done_count)
                    if refined_diff < current_diff and refine_meta:
                        print(
                            "Применен доп.срез к UI-счетчику: "
                            f"{refine_meta} (UI={ui_done_count})"
                        )
                        orders = refined_orders

            if args.max_orders is not None:
                orders = orders[: args.max_orders]

            print(f"Найдено заказов: {len(orders)}")

            all_statuses: list[dict[str, Any]] = []
            for index, order in enumerate(orders, start=1):
                sale_id = order.get("Sale")
                number = order.get("Number")
                print(f"[history] {index}/{len(orders)} sale={sale_id} number={number}")
                statuses = fetch_order_status_history(
                    client=client,
                    template=templates,
                    order=order,
                    history_page_limit=int(args.history_page_limit),
                    max_history_pages=args.max_history_pages,
                )
                all_statuses.extend(statuses)

            print(f"Найдено смен статуса: {len(all_statuses)}")
            export_excel(output_path, target_date, orders, all_statuses)
            print(f"Excel сохранен: {output_path}")

        finally:
            context.close()

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("Остановлено пользователем")
        raise SystemExit(130)
